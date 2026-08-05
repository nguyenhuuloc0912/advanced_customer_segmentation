"""
Customer Segmentation Dashboard
Streamlit app for interactive exploration of customer clustering results.

Expected files in data/processed/:
  - cleaned_uk_data.csv
  - customer_features.csv
  - customer_features_scaled.csv
  - customer_clusters_k3.csv
  - customer_clusters_k4.csv
"""

import os
import warnings
import io
from datetime import datetime

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from plotly.subplots import make_subplots
from scipy import stats as scipy_stats
from sklearn.cluster import KMeans
from sklearn.decomposition import PCA
from sklearn.preprocessing import StandardScaler

# ─────────────────────────────────────────────
# PDF EXPORT DEPENDENCIES (graceful fallback nếu chưa cài đặt)
# ─────────────────────────────────────────────
try:
    import matplotlib
    matplotlib.use("Agg")  # render không cần màn hình, an toàn khi chạy trên server
    import matplotlib.pyplot as plt
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Image as RLImage,
        Table, TableStyle, PageBreak, HRFlowable,
    )

    # DejaVu Sans được đóng gói sẵn bên trong matplotlib, hỗ trợ đầy đủ dấu tiếng Việt,
    # nên không cần thêm file font riêng vào project.
    _FONT_DIR = os.path.join(matplotlib.get_data_path(), "fonts", "ttf")
    pdfmetrics.registerFont(TTFont("VNSans", os.path.join(_FONT_DIR, "DejaVuSans.ttf")))
    pdfmetrics.registerFont(TTFont("VNSans-Bold", os.path.join(_FONT_DIR, "DejaVuSans-Bold.ttf")))
    PDF_EXPORT_AVAILABLE = True
except Exception as _pdf_import_error:
    PDF_EXPORT_AVAILABLE = False
    _PDF_IMPORT_ERROR_MSG = str(_pdf_import_error)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter

    EXCEL_EXPORT_AVAILABLE = True
except Exception as _xlsx_import_error:
    EXCEL_EXPORT_AVAILABLE = False
    _XLSX_IMPORT_ERROR_MSG = str(_xlsx_import_error)

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="Customer Segmentation Dashboard",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────
# GLOBAL STYLE
# ─────────────────────────────────────────────
st.markdown(
    """
<style>
  @import url('https://fonts.googleapis.com/css2?family=Space+Mono:wght@400;700&family=DM+Sans:wght@300;400;500;600&display=swap');

  :root {
    --bg: #0a0d14;
    --surface: #11151f;
    --card: #171c28;
    --border: #262c3a;
    --accent1: #7c8aff;
    --accent2: #ff8fa3;
    --accent3: #2dd4a7;
    --accent4: #f0b429;
    --accent5: #a78bfa;
    --accent6: #7b8794;
    --text: #edf0f5;
    --muted: #8894a8;
  }

  html, body, [data-testid="stAppViewContainer"] {
        background:
            radial-gradient(circle at top left, rgba(124, 138, 255, 0.16), transparent 30%),
            radial-gradient(circle at top right, rgba(240, 180, 41, 0.10), transparent 26%),
            linear-gradient(180deg, #0a0d14 0%, #080a10 100%) !important;
    color: var(--text) !important;
    font-family: 'DM Sans', sans-serif;
  }

    [data-testid="stAppViewContainer"]::before {
        content: '';
        position: fixed;
        inset: 0;
        pointer-events: none;
        background-image: linear-gradient(rgba(255,255,255,0.02) 1px, transparent 1px), linear-gradient(90deg, rgba(255,255,255,0.02) 1px, transparent 1px);
        background-size: 28px 28px;
        mask-image: linear-gradient(180deg, rgba(0,0,0,0.45), transparent 85%);
        opacity: 0.35;
    }

  [data-testid="stSidebar"] {
    background-color: var(--surface) !important;
    border-right: 1px solid var(--border);
  }

  h1, h2, h3 { font-family: 'Space Mono', monospace !important; }

  .metric-card {
    background: var(--card);
    border: 1px solid var(--border);
    border-radius: 12px;
    padding: 18px 22px;
    margin-bottom: 12px;
    transition: border-color 0.2s;
        box-shadow: 0 18px 40px rgba(0, 0, 0, 0.22);
  }
  .metric-card:hover { border-color: var(--accent1); }

    .hero-banner {
        background: linear-gradient(135deg, rgba(108, 99, 255, 0.18), rgba(255, 101, 132, 0.10));
        border: 1px solid rgba(255, 255, 255, 0.08);
        border-radius: 18px;
        padding: 22px 24px;
        margin: 4px 0 18px 0;
        box-shadow: 0 22px 46px rgba(0, 0, 0, 0.18);
    }
    .hero-kicker {
        display: inline-block;
        font-family: 'Space Mono', monospace;
        font-size: 0.72rem;
        letter-spacing: 0.18em;
        text-transform: uppercase;
        color: var(--accent3);
        margin-bottom: 8px;
    }
    .hero-title {
        font-family: 'Space Mono', monospace;
        font-size: 2rem;
        font-weight: 700;
        line-height: 1.2;
        margin-bottom: 8px;
    }
    .hero-subtitle {
        color: var(--muted);
        font-size: 0.98rem;
        line-height: 1.6;
        max-width: 860px;
    }
    .hero-chips {
        display: flex;
        flex-wrap: wrap;
        gap: 8px;
        margin-top: 14px;
    }
    .hero-chip {
        display: inline-flex;
        align-items: center;
        gap: 6px;
        background: rgba(255,255,255,0.04);
        border: 1px solid var(--border);
        border-radius: 999px;
        padding: 6px 12px;
        font-size: 0.82rem;
        color: var(--text);
    }

  .metric-value {
    font-family: 'Space Mono', monospace;
    font-size: 2rem;
    font-weight: 700;
    line-height: 1.1;
  }
  .metric-label {
    font-size: 0.78rem;
    color: var(--muted);
    text-transform: uppercase;
    letter-spacing: 0.08em;
    margin-top: 4px;
  }

  .cluster-badge {
    display: inline-block;
    padding: 3px 10px;
    border-radius: 20px;
    font-size: 0.75rem;
    font-weight: 600;
    font-family: 'Space Mono', monospace;
  }

  .section-header {
    font-family: 'Space Mono', monospace;
    font-size: 0.72rem;
    letter-spacing: 0.15em;
    text-transform: uppercase;
    color: var(--muted);
    margin-bottom: 16px;
    padding-bottom: 8px;
    border-bottom: 1px solid var(--border);
  }

  .stSelectbox > div, .stSlider > div {
    background: var(--card) !important;
  }

  div[data-testid="metric-container"] {
    background: var(--card);
    border: 1px solid var(--border);
    border-radius: 10px;
    padding: 14px 18px;
  }

  .stTabs [data-baseweb="tab-list"] {
    background: var(--surface);
    border-bottom: 1px solid var(--border);
    gap: 4px;
  }
  .stTabs [data-baseweb="tab"] {
    background: transparent;
    color: var(--muted);
    font-family: 'Space Mono', monospace;
    font-size: 0.78rem;
    border-radius: 6px 6px 0 0;
    padding: 8px 16px;
  }
  .stTabs [aria-selected="true"] {
    background: var(--card) !important;
    color: var(--accent1) !important;
    border-bottom: 2px solid var(--accent1);
  }
</style>
""",
    unsafe_allow_html=True,
)

# ─────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────
CLUSTER_COLORS_K3 = {0: "#f0b429", 1: "#2dd4a7", 2: "#a78bfa"}
CLUSTER_COLORS_K4 = {0: "#7b8794", 1: "#f0b429", 2: "#2dd4a7", 3: "#a78bfa"}

FEATURE_NAMES_VN = {
    "Sum_Quantity": "Tổng số lượng mua",
    "Mean_UnitPrice": "Giá trung bình",
    "Mean_TotalPrice": "Giá trị giao dịch TB",
    "Sum_TotalPrice": "Tổng chi tiêu",
    "Count_Invoice": "Số lần mua",
    "Count_Stock": "Số sản phẩm khác nhau",
    "Mean_InvoiceCountPerStock": "Tần suất mua/sản phẩm",
    "Mean_StockCountPerInvoice": "Sản phẩm/giao dịch",
    "Mean_UnitPriceMeanPerInvoice": "Giá TB/giao dịch",
    "Mean_QuantitySumPerInvoice": "Số lượng/giao dịch",
    "Mean_TotalPriceMeanPerInvoice": "Giá trị TB/giao dịch",
    "Mean_TotalPriceSumPerInvoice": "Tổng giá trị/giao dịch",
    "Mean_UnitPriceMeanPerStock": "Giá TB/sản phẩm",
    "Mean_QuantitySumPerStock": "Số lượng TB/sản phẩm",
    "Mean_TotalPriceMeanPerStock": "Giá trị TB/sản phẩm",
    "Mean_TotalPriceSumPerStock": "Tổng giá trị/sản phẩm",
}

RADAR_FEATURES = {
    "Sum_Quantity": "Khối lượng mua",
    "Sum_TotalPrice": "Tổng chi tiêu",
    "Mean_UnitPrice": "Mức giá ưa thích",
    "Count_Invoice": "Tần suất mua",
    "Count_Stock": "Đa dạng sản phẩm",
    "Mean_TotalPriceSumPerInvoice": "Giá trị/giao dịch",
    "Mean_TotalPriceMeanPerStock": "Chi tiêu/sản phẩm",
    "Mean_StockCountPerInvoice": "Sản phẩm/giao dịch",
}

SEGMENT_PERSONAS_K4 = {
    0: {"name": "Nhóm giá trị thấp", "color": "#7b8794", "desc": "Giá mua, tổng chi tiêu và tần suất đều thấp nhất — cơ hội để kích hoạt lại"},
    1: {"name": "Nhóm Premium", "color": "#f0b429", "desc": "Giá mua trung bình cao nhất, mua ít nhưng chi mạnh cho từng sản phẩm"},
    2: {"name": "Nhóm mua nhiều thường xuyên", "color": "#2dd4a7", "desc": "Tổng chi tiêu, số lượng và tần suất mua đều cao nhất"},
    3: {"name": "Nhóm đa dạng sản phẩm", "color": "#a78bfa", "desc": "Danh mục mua sắm rộng và phân tán hơn"},
}

SEGMENT_PERSONAS_K3 = {
    0: {"name": "Nhóm giá trị cao", "color": "#f0b429", "desc": "Chi tiêu và mức giá mua cao"},
    1: {"name": "Nhóm mua nhiều", "color": "#2dd4a7", "desc": "Tổng chi tiêu và khối lượng mua lớn"},
    2: {"name": "Nhóm đa dạng", "color": "#a78bfa", "desc": "Khách hàng có danh mục mua sắm đa dạng"},
}

MARKETING_STRATEGY_K4 = {
    0: "Ưu đãi kích hoạt lại (win-back), khảo sát nguyên nhân giảm tương tác, nhắc nhở qua email/SMS định kỳ.",
    1: "Chương trình khách hàng VIP, quyền tiếp cận sớm sản phẩm mới, trải nghiệm cao cấp — hạn chế giảm giá trực tiếp để không làm giảm giá trị cảm nhận.",
    2: "Chương trình tích điểm theo cấp bậc, ưu đãi theo số lượng mua, cân nhắc mô hình đăng ký (subscription) định kỳ.",
    3: "Xây dựng hệ thống gợi ý sản phẩm (recommendation engine) và chương trình cross-selling cá nhân hóa dựa trên lịch sử mua đa dạng.",
}
MARKETING_STRATEGY_K3 = {
    0: "Chương trình khách hàng thân thiết, ưu đãi theo giá trị đơn hàng, dịch vụ chăm sóc cá nhân hóa.",
    1: "Chính sách chiết khấu theo số lượng, ưu tiên giao hàng nhanh cho đơn hàng lớn.",
    2: "Gợi ý sản phẩm đa dạng, chương trình khuyến khích khám phá danh mục sản phẩm mới.",
}

DATA_DIR = "data/processed"
MERGED_FILES = {
    "cleaned": f"{DATA_DIR}/cleaned_uk_data_merged.csv",
    "features": f"{DATA_DIR}/customer_features_merged.csv",
    "scaled": f"{DATA_DIR}/customer_features_scaled_merged.csv",
    "clusters_k3": f"{DATA_DIR}/customer_clusters_k3_merged.csv",
    "clusters_k4": f"{DATA_DIR}/customer_clusters_k4_merged.csv",
}


def save_merged_to_disk(rec):
    """Ghi dữ liệu đã gộp/tính lại ra các file *_merged.csv riêng biệt trong data/processed/,
    KHÔNG đè lên các file gốc (cleaned_uk_data.csv, customer_features.csv...).
    Nhờ đó dữ liệu vẫn còn sau khi tải lại trang (F5) hoặc khởi động lại app,
    trong khi vẫn giữ nguyên baseline gốc để có thể "Khôi phục dữ liệu gốc" bất kỳ lúc nào."""
    try:
        os.makedirs(DATA_DIR, exist_ok=True)
        rec["cleaned"].to_csv(MERGED_FILES["cleaned"], index=False)
        rec["features"].to_csv(MERGED_FILES["features"], index=True)
        rec["scaled"].to_csv(MERGED_FILES["scaled"], index=True)
        rec["clusters_k3"].to_csv(MERGED_FILES["clusters_k3"], index=False)
        rec["clusters_k4"].to_csv(MERGED_FILES["clusters_k4"], index=False)
        return True, None
    except Exception as e:
        return False, str(e)


def load_merged_from_disk():
    """Đọc lại dữ liệu đã gộp từ các file *_merged.csv nếu tồn tại. Trả về None nếu chưa từng lưu."""
    if not all(os.path.exists(p) for p in MERGED_FILES.values()):
        return None
    try:
        cleaned = pd.read_csv(MERGED_FILES["cleaned"])
        features = pd.read_csv(MERGED_FILES["features"], index_col=0)
        scaled = pd.read_csv(MERGED_FILES["scaled"], index_col=0)
        clusters_k3 = pd.read_csv(MERGED_FILES["clusters_k3"])
        clusters_k4 = pd.read_csv(MERGED_FILES["clusters_k4"])
        features.index = features.index.astype(str)
        scaled.index = scaled.index.astype(str)
        n_total = len(features)
        return {
            "cleaned": cleaned, "features": features, "scaled": scaled,
            "clusters_k3": clusters_k3, "clusters_k4": clusters_k4,
            "n_new_customers": None,  # không rõ số khách mới sau khi đọc lại từ đĩa
            "n_total_customers": n_total,
        }
    except Exception:
        return None


def delete_merged_from_disk():
    for p in MERGED_FILES.values():
        if os.path.exists(p):
            try:
                os.remove(p)
            except Exception:
                pass


# ─────────────────────────────────────────────
# DATA LOADING
# ─────────────────────────────────────────────
@st.cache_data
def load_all_data():
    data = {}
    files = {
        "cleaned": f"{DATA_DIR}/cleaned_uk_data.csv",
        "features": f"{DATA_DIR}/customer_features.csv",
        "scaled": f"{DATA_DIR}/customer_features_scaled.csv",
        "clusters_k3": f"{DATA_DIR}/customer_clusters_k3.csv",
        "clusters_k4": f"{DATA_DIR}/customer_clusters_k4.csv",
    }
    for key, path in files.items():
        if os.path.exists(path):
            data[key] = pd.read_csv(path, index_col=0 if key in ("features", "scaled") else None)
        else:
            data[key] = None
    return data


def get_cluster_colors(k):
    return CLUSTER_COLORS_K4 if k == 4 else CLUSTER_COLORS_K3


def get_personas(k):
    return SEGMENT_PERSONAS_K4 if k == 4 else SEGMENT_PERSONAS_K3


def merge_clusters(features_df, clusters_df, k):
    if features_df is None or clusters_df is None:
        return None
    col = "Cluster"
    clusters_df = clusters_df.copy()
    clusters_df["CustomerID"] = clusters_df["CustomerID"].astype(str)
    features = features_df.copy()
    features.index = features.index.astype(str)
    merged = features.merge(clusters_df.set_index("CustomerID")[col], left_index=True, right_index=True, how="left")
    merged.rename(columns={col: f"Cluster_{k}"}, inplace=True)
    return merged


# ─────────────────────────────────────────────
# UPLOAD & RECOMPUTE PIPELINE
# ─────────────────────────────────────────────
REQUIRED_RAW_COLUMNS = [
    "InvoiceNo", "StockCode", "Description", "Quantity",
    "InvoiceDate", "UnitPrice", "CustomerID", "Country",
]


def normalize_customer_id(series):
    """Chuẩn hóa CustomerID về CHUỖI SỐ NGUYÊN nhất quán, bất kể dữ liệu gốc
    lưu dưới dạng float (17850.0), int (17850), hay string ("17850", "17850.0").
    Bắt buộc phải gọi hàm này ở MỌI nơi dữ liệu cũ và dữ liệu mới giao nhau,
    nếu không groupby/merge theo CustomerID sẽ đếm trùng khách hàng do lệch kiểu dữ liệu
    (vd: 17850.0 và "17850" bị coi là hai khách hàng khác nhau)."""
    numeric = pd.to_numeric(series, errors="coerce")
    return numeric.round().astype("Int64").astype(str)


def clean_new_upload(df_raw):
    """Áp dụng đúng các bước làm sạch đã dùng trong notebook 01_cleaning_and_eda
    lên dữ liệu giao dịch mới do doanh nghiệp tải lên."""
    df = df_raw.copy()
    missing = [c for c in REQUIRED_RAW_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(
            "File thiếu các cột bắt buộc: " + ", ".join(missing) +
            ". Cần đúng cấu trúc: " + ", ".join(REQUIRED_RAW_COLUMNS)
        )

    df = df.dropna(subset=["CustomerID"])
    df["InvoiceNo"] = df["InvoiceNo"].astype(str)
    df = df[~df["InvoiceNo"].str.startswith("C")]  # loại hóa đơn hủy
    df = df[df["Country"].astype(str).str.strip() == "United Kingdom"]  # chỉ thị trường UK
    df = df[(df["Quantity"] > 0) & (df["UnitPrice"] > 0)]  # loại số lượng/giá không hợp lệ

    df["CustomerID"] = normalize_customer_id(df["CustomerID"])
    df["InvoiceDate"] = pd.to_datetime(df["InvoiceDate"], errors="coerce")
    return df.reset_index(drop=True)


def build_customer_features(df_clean):
    """Xây dựng lại đúng 16 đặc trưng khách hàng (4 nhóm) từ dữ liệu giao dịch đã làm sạch,
    theo cùng định nghĩa đã dùng trong notebook 02_feature_engineering."""
    df = df_clean.copy()
    df["TotalPrice"] = df["Quantity"] * df["UnitPrice"]

    # Nhóm cơ bản (6)
    basic = df.groupby("CustomerID").agg(
        Sum_Quantity=("Quantity", "sum"),
        Mean_UnitPrice=("UnitPrice", "mean"),
        Mean_TotalPrice=("TotalPrice", "mean"),
        Sum_TotalPrice=("TotalPrice", "sum"),
        Count_Invoice=("InvoiceNo", "nunique"),
        Count_Stock=("StockCode", "nunique"),
    )

    # Nhóm theo sản phẩm (2)
    stock_invoice = df.groupby(["CustomerID", "StockCode"])["InvoiceNo"].nunique()
    mean_invoice_count_per_stock = stock_invoice.groupby("CustomerID").mean()

    invoice_stock = df.groupby(["CustomerID", "InvoiceNo"])["StockCode"].nunique()
    mean_stock_count_per_invoice = invoice_stock.groupby("CustomerID").mean()

    # Nhóm theo hóa đơn (4)
    per_invoice = df.groupby(["CustomerID", "InvoiceNo"]).agg(
        UnitPriceMean=("UnitPrice", "mean"),
        QuantitySum=("Quantity", "sum"),
        TotalPriceMean=("TotalPrice", "mean"),
        TotalPriceSum=("TotalPrice", "sum"),
    )
    invoice_based = per_invoice.groupby("CustomerID").agg(
        Mean_UnitPriceMeanPerInvoice=("UnitPriceMean", "mean"),
        Mean_QuantitySumPerInvoice=("QuantitySum", "mean"),
        Mean_TotalPriceMeanPerInvoice=("TotalPriceMean", "mean"),
        Mean_TotalPriceSumPerInvoice=("TotalPriceSum", "mean"),
    )

    # Nhóm theo loại sản phẩm (4)
    per_stock = df.groupby(["CustomerID", "StockCode"]).agg(
        UnitPriceMean=("UnitPrice", "mean"),
        QuantitySum=("Quantity", "sum"),
        TotalPriceMean=("TotalPrice", "mean"),
        TotalPriceSum=("TotalPrice", "sum"),
    )
    stock_based = per_stock.groupby("CustomerID").agg(
        Mean_UnitPriceMeanPerStock=("UnitPriceMean", "mean"),
        Mean_QuantitySumPerStock=("QuantitySum", "mean"),
        Mean_TotalPriceMeanPerStock=("TotalPriceMean", "mean"),
        Mean_TotalPriceSumPerStock=("TotalPriceSum", "mean"),
    )

    features = basic.copy()
    features["Mean_InvoiceCountPerStock"] = mean_invoice_count_per_stock
    features["Mean_StockCountPerInvoice"] = mean_stock_count_per_invoice
    features = features.join(invoice_based).join(stock_based)
    features.index = features.index.astype(str)
    features.index.name = "CustomerID"
    return features[[
        "Sum_Quantity", "Mean_UnitPrice", "Mean_TotalPrice", "Sum_TotalPrice", "Count_Invoice", "Count_Stock",
        "Mean_InvoiceCountPerStock", "Mean_StockCountPerInvoice",
        "Mean_UnitPriceMeanPerInvoice", "Mean_QuantitySumPerInvoice",
        "Mean_TotalPriceMeanPerInvoice", "Mean_TotalPriceSumPerInvoice",
        "Mean_UnitPriceMeanPerStock", "Mean_QuantitySumPerStock",
        "Mean_TotalPriceMeanPerStock", "Mean_TotalPriceSumPerStock",
    ]]


def transform_and_cluster(features_df, k_list=(3, 4), random_state=42):
    """Box-Cox -> StandardScaler -> PCA (>=85% phương sai) -> KMeans(k) cho từng k trong k_list."""
    X = features_df.copy()

    boxcox_df = pd.DataFrame(index=X.index)
    for col in X.columns:
        vals = X[col].values.astype(float)
        shift = abs(vals.min()) + 1.0 if vals.min() <= 0 else 0.0
        try:
            transformed, _ = scipy_stats.boxcox(vals + shift)
        except Exception:
            transformed = np.log1p(vals + shift)  # fallback nếu Box-Cox không hội tụ
        boxcox_df[col] = transformed

    scaler = StandardScaler()
    scaled_vals = scaler.fit_transform(boxcox_df.values)
    scaled_df = pd.DataFrame(scaled_vals, index=X.index, columns=X.columns)

    pca = PCA(n_components=0.85, svd_solver="full", random_state=random_state)
    pca_coords = pca.fit_transform(scaled_vals)

    cluster_results = {}
    for k in k_list:
        km = KMeans(n_clusters=k, random_state=random_state, n_init=10)
        labels = km.fit_predict(pca_coords)
        cluster_results[k] = pd.DataFrame({"CustomerID": X.index, "Cluster": labels})

    return scaled_df, pca_coords, pca.explained_variance_ratio_, cluster_results


# ─────────────────────────────────────────────
# PDF REPORT BUILDER
# ─────────────────────────────────────────────
def _fig_to_rlimage(fig, width_cm=14, height_cm=8):
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return RLImage(buf, width=width_cm * cm, height=height_cm * cm)


def build_pdf_report(cleaned_df, features_df, clusters_k3, clusters_k4, data_label):
    """Tạo báo cáo PDF tổng hợp: tổng quan dữ liệu, phân bố phân khúc (k=3, k=4),
    đặc trưng hành vi từng nhóm (radar), và đề xuất chiến lược marketing.
    Dùng font DejaVu (đóng gói sẵn trong matplotlib) để hỗ trợ đầy đủ tiếng Việt có dấu."""
    font_regular, font_bold = "VNSans", "VNSans-Bold"

    style_title = ParagraphStyle("title", fontName=font_bold, fontSize=24, leading=30,
                                  textColor=rl_colors.HexColor("#1a1a2e"), alignment=TA_CENTER, spaceAfter=6)
    style_subtitle = ParagraphStyle("subtitle", fontName=font_regular, fontSize=13, leading=18,
                                     textColor=rl_colors.HexColor("#5f5e5a"), alignment=TA_CENTER, spaceAfter=4)
    style_h2 = ParagraphStyle("h2", fontName=font_bold, fontSize=16, leading=20,
                               textColor=rl_colors.HexColor("#1a1a2e"), spaceBefore=10, spaceAfter=8)
    style_h3 = ParagraphStyle("h3", fontName=font_bold, fontSize=12.5, spaceBefore=6, spaceAfter=6)
    style_seg_title = ParagraphStyle("seg_title", fontName=font_bold, fontSize=12, spaceBefore=10, spaceAfter=2)
    style_body = ParagraphStyle("body", fontName=font_regular, fontSize=10.5, leading=15,
                                 textColor=rl_colors.HexColor("#2c2c2a"), spaceAfter=6)
    style_caption = ParagraphStyle("caption", fontName=font_regular, fontSize=8.5, leading=12,
                                    textColor=rl_colors.HexColor("#888780"))

    story = []

    # ── Trang bìa ──
    story.append(Spacer(1, 5 * cm))
    story.append(Paragraph("BÁO CÁO PHÂN KHÚC KHÁCH HÀNG", style_title))
    story.append(Paragraph("Tổng hợp kết quả phân tích dữ liệu bằng học máy không giám sát", style_subtitle))
    story.append(Spacer(1, 1 * cm))
    story.append(HRFlowable(width="40%", thickness=1, color=rl_colors.HexColor("#e8871e"), hAlign="CENTER"))
    story.append(Spacer(1, 1 * cm))
    story.append(Paragraph(f"Ngày xuất báo cáo: {datetime.now().strftime('%d/%m/%Y %H:%M')}", style_subtitle))
    story.append(Paragraph(f"Nguồn dữ liệu: {data_label}", style_subtitle))
    story.append(PageBreak())

    # ── 1. Tổng quan dữ liệu ──
    story.append(Paragraph("1. Tổng quan dữ liệu", style_h2))
    n_customers = len(features_df)
    n_transactions = len(cleaned_df) if cleaned_df is not None else None
    total_revenue = features_df["Sum_TotalPrice"].sum() if "Sum_TotalPrice" in features_df.columns else None
    avg_spend = features_df["Sum_TotalPrice"].mean() if "Sum_TotalPrice" in features_df.columns else None
    avg_txn = features_df["Count_Invoice"].mean() if "Count_Invoice" in features_df.columns else None

    kpi_rows = [["Chỉ số", "Giá trị"], ["Tổng số khách hàng", f"{n_customers:,}"]]
    if n_transactions is not None:
        kpi_rows.append(["Tổng số giao dịch", f"{n_transactions:,}"])
    if total_revenue is not None:
        kpi_rows.append(["Tổng doanh thu", f"£{total_revenue:,.0f}"])
    if avg_spend is not None:
        kpi_rows.append(["Chi tiêu trung bình / khách hàng", f"£{avg_spend:,.2f}"])
    if avg_txn is not None:
        kpi_rows.append(["Số giao dịch trung bình / khách hàng", f"{avg_txn:,.1f}"])

    kpi_table = Table(kpi_rows, colWidths=[9 * cm, 6 * cm])
    kpi_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_regular),
        ("FONTNAME", (0, 0), (-1, 0), font_bold),
        ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#1e2761")),
        ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, rl_colors.HexColor("#d8d8d0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor("#f7f8fc")]),
        ("FONTSIZE", (0, 0), (-1, -1), 10.5),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 0.5 * cm))
    story.append(Paragraph(
        "Số liệu phản ánh dữ liệu hiện tại đang được sử dụng trong dashboard tại thời điểm xuất báo cáo.",
        style_caption
    ))
    story.append(PageBreak())

    # ── 2. Phân bố khách hàng theo phân khúc (k=4 và k=3) ──
    story.append(Paragraph("2. Phân bố khách hàng theo phân khúc", style_h2))

    for k, clusters_df, personas in [(4, clusters_k4, SEGMENT_PERSONAS_K4), (3, clusters_k3, SEGMENT_PERSONAS_K3)]:
        if clusters_df is None or len(clusters_df) == 0:
            continue
        story.append(Paragraph(f"Phương án k = {k}", style_h3))
        counts = clusters_df["Cluster"].value_counts().sort_index()
        total = counts.sum()
        labels = [personas.get(int(c), {}).get("name", f"Cụm {c}") for c in counts.index]
        colors_list = [personas.get(int(c), {}).get("color", "#888888") for c in counts.index]

        fig, ax = plt.subplots(figsize=(5.5, 4))
        wedges, _ = ax.pie(counts.values, colors=colors_list, startangle=90,
                            wedgeprops=dict(width=0.42, edgecolor="white"))
        ax.legend(wedges, [f"{l} ({c / total:.1%})" for l, c in zip(labels, counts.values)],
                  loc="center left", bbox_to_anchor=(1.0, 0.5), fontsize=9, frameon=False)
        ax.set_title(f"Phân bố khách hàng — k={k}", fontsize=12)
        story.append(_fig_to_rlimage(fig, width_cm=14, height_cm=8))

        table_rows = [["Phân khúc", "Số khách hàng", "Tỷ lệ"]]
        for label, cnt in zip(labels, counts.values):
            table_rows.append([label, f"{cnt:,}", f"{cnt / total:.1%}"])
        t = Table(table_rows, colWidths=[7 * cm, 4 * cm, 3 * cm])
        t.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font_regular),
            ("FONTNAME", (0, 0), (-1, 0), font_bold),
            ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#1e2761")),
            ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, rl_colors.HexColor("#d8d8d0")),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.6 * cm))
    story.append(PageBreak())

    # ── 3. Đặc trưng hành vi từng phân khúc (radar, k=4) ──
    story.append(Paragraph("3. Đặc trưng hành vi từng phân khúc (k = 4)", style_h2))
    if clusters_k4 is not None and features_df is not None and len(clusters_k4) > 0:
        merged4 = features_df.copy()
        merged4.index = merged4.index.astype(str)
        ck4 = clusters_k4.copy()
        ck4["CustomerID"] = ck4["CustomerID"].astype(str)
        merged4 = merged4.merge(ck4.set_index("CustomerID")["Cluster"], left_index=True, right_index=True, how="left")

        radar_keys = [f for f in RADAR_FEATURES.keys() if f in merged4.columns]
        cluster_means = merged4.groupby("Cluster")[radar_keys].mean()
        gmin, gmax = cluster_means.min(), cluster_means.max()
        norm = (cluster_means - gmin) / (gmax - gmin + 1e-9)
        categories = [RADAR_FEATURES[f] for f in radar_keys]

        angles = np.linspace(0, 2 * np.pi, len(categories), endpoint=False).tolist()
        angles += angles[:1]

        fig, ax = plt.subplots(figsize=(6, 6), subplot_kw=dict(polar=True))
        for cid, row in norm.iterrows():
            vals = row.tolist() + row.tolist()[:1]
            color = SEGMENT_PERSONAS_K4.get(int(cid), {}).get("color", "#888888")
            name = SEGMENT_PERSONAS_K4.get(int(cid), {}).get("name", f"Cụm {cid}")
            ax.plot(angles, vals, color=color, linewidth=1.8, label=name)
            ax.fill(angles, vals, color=color, alpha=0.12)
        ax.set_xticks(angles[:-1])
        ax.set_xticklabels(categories, fontsize=8.5)
        ax.set_yticklabels([])
        ax.legend(loc="upper right", bbox_to_anchor=(1.4, 1.12), fontsize=8.5, frameon=False)
        ax.set_title("So sánh đặc trưng trung bình từng phân khúc (đã chuẩn hóa)", fontsize=11, pad=20)
        story.append(_fig_to_rlimage(fig, width_cm=13, height_cm=13))
    else:
        story.append(Paragraph("Không có đủ dữ liệu để dựng biểu đồ radar.", style_body))
    story.append(PageBreak())

    # ── 4. Đề xuất chiến lược marketing ──
    story.append(Paragraph("4. Đề xuất chiến lược marketing theo phân khúc", style_h2))
    for cid in sorted(SEGMENT_PERSONAS_K4.keys()):
        persona = SEGMENT_PERSONAS_K4[cid]
        story.append(Paragraph(
            f'<font color="{persona["color"]}">●</font> <b>{persona["name"]}</b>', style_seg_title
        ))
        story.append(Paragraph(persona["desc"], style_body))
        story.append(Paragraph(f'<b>Đề xuất:</b> {MARKETING_STRATEGY_K4.get(cid, "")}', style_body))

    story.append(Spacer(1, 1 * cm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=rl_colors.HexColor("#d8d8d0")))
    story.append(Spacer(1, 0.3 * cm))
    story.append(Paragraph(
        f"Báo cáo được tạo tự động từ Dashboard Phân Khúc Khách Hàng — "
        f"{datetime.now().strftime('%d/%m/%Y %H:%M')}. Dữ liệu: {n_customers:,} khách hàng ({data_label}).",
        style_caption
    ))

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=2 * cm, bottomMargin=2 * cm,
                             leftMargin=2 * cm, rightMargin=2 * cm)
    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────
# EXCEL DASHBOARD BUILDER (kiểu Power BI: dropdown lọc + biểu đồ + bảng có AutoFilter)
# ─────────────────────────────────────────────
HEADER_FILL = "1E2761"
ACCENT_FILL = "E8871E"
LIGHT_FILL = "F7F8FC"


def _style_header_row(ws, row, n_cols, fill=HEADER_FILL):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def build_excel_dashboard(cleaned_df, features_df, clusters_k3, clusters_k4, data_label):
    """Tạo file Excel dạng dashboard: sheet Dashboard (KPI + dropdown lọc nhanh + biểu đồ),
    sheet Dữ Liệu Khách Hàng (Excel Table có AutoFilter từng cột), sheet So Sánh Cụm (color scale).

    Lưu ý: đây KHÔNG phải Slicer thật của Excel/Power BI (không có thư viện Python nào tạo được
    Slicer từ đầu) — mà là ô dropdown (Data Validation) kết hợp công thức SUMIF/COUNTIF/AVERAGEIF,
    cho hiệu quả lọc tương tự nhưng giao diện khác."""
    df = features_df.copy()
    df.index = df.index.astype(str)
    df.index.name = "CustomerID"
    df = df.reset_index()

    for k, clusters_df in [(3, clusters_k3), (4, clusters_k4)]:
        personas = SEGMENT_PERSONAS_K4 if k == 4 else SEGMENT_PERSONAS_K3
        col_id, col_name = f"Cluster_k{k}", f"Cluster_k{k}_Ten"
        if clusters_df is not None and len(clusters_df) > 0:
            cdf = clusters_df.copy()
            cdf["CustomerID"] = cdf["CustomerID"].astype(str)
            cdf = cdf.rename(columns={"Cluster": col_id})
            df = df.merge(cdf[["CustomerID", col_id]], on="CustomerID", how="left")
            df[col_name] = df[col_id].map(
                lambda c: personas.get(int(c), {}).get("name", f"Cụm {int(c)}") if pd.notna(c) else ""
            )
        else:
            df[col_id] = None
            df[col_name] = ""

    wb = Workbook()

    # ══ Sheet "Dữ Liệu Khách Hàng" ══
    ws_data = wb.active
    ws_data.title = "Dữ Liệu Khách Hàng"

    display_cols = ["CustomerID"] + list(features_df.columns) + ["Cluster_k4_Ten", "Cluster_k3_Ten"]
    header_labels = ["Mã khách hàng"] + [FEATURE_NAMES_VN.get(c, c) for c in features_df.columns] + \
                     ["Phân khúc (k=4)", "Phân khúc (k=3)"]

    ws_data.append(header_labels)
    for _, row in df[display_cols].iterrows():
        ws_data.append(list(row))

    n_rows, n_cols = len(df) + 1, len(display_cols)
    _style_header_row(ws_data, 1, n_cols)

    for col_idx in range(1, n_cols + 1):
        ws_data.column_dimensions[get_column_letter(col_idx)].width = 20 if col_idx > 1 else 16
    ws_data.freeze_panes = "B2"

    table_ref = f"A1:{get_column_letter(n_cols)}{n_rows}"
    tbl = Table(displayName="TblData", ref=table_ref)
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws_data.add_table(tbl)

    # ══ Sheet "Dashboard" ══
    ws_dash = wb.create_sheet("Dashboard", 0)
    ws_dash.sheet_view.showGridLines = False

    ws_dash.merge_cells("B2:H2")
    ws_dash["B2"] = "DASHBOARD PHÂN KHÚC KHÁCH HÀNG"
    ws_dash["B2"].font = Font(name="Calibri", bold=True, size=20, color=HEADER_FILL)

    ws_dash.merge_cells("B3:H3")
    ws_dash["B3"] = f"Ngày xuất: {datetime.now().strftime('%d/%m/%Y %H:%M')} · Nguồn dữ liệu: {data_label}"
    ws_dash["B3"].font = Font(name="Calibri", italic=True, size=10.5, color="6B7A99")

    ws_dash["B5"] = "🔍 Bộ lọc nhanh — chọn phân khúc (k=4):"
    ws_dash["B5"].font = Font(name="Calibri", bold=True, size=11)
    persona_names = [SEGMENT_PERSONAS_K4[c]["name"] for c in sorted(SEGMENT_PERSONAS_K4.keys())]
    ws_dash["E5"] = "Tất cả"
    ws_dash["E5"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws_dash["E5"].fill = PatternFill("solid", fgColor=ACCENT_FILL)
    ws_dash["E5"].alignment = Alignment(horizontal="center")
    dv_list = '"' + ",".join(["Tất cả"] + persona_names) + '"'
    dv = DataValidation(type="list", formula1=dv_list, allow_blank=False)
    ws_dash.add_data_validation(dv)
    dv.add(ws_dash["E5"])

    kpi_defs = [
        ("Số khách hàng",
         '=IF($E$5="Tất cả",COUNTA(TblData[Mã khách hàng]),COUNTIF(TblData[Phân khúc (k=4)],$E$5))', "#,##0"),
        ("Tổng chi tiêu (£)",
         '=IF($E$5="Tất cả",SUM(TblData[Tổng chi tiêu]),SUMIF(TblData[Phân khúc (k=4)],$E$5,TblData[Tổng chi tiêu]))', "#,##0"),
        ("Chi tiêu TB / khách (£)", "=C8/C7", "#,##0.00"),
        ("Số lần mua TB / khách",
         '=IF($E$5="Tất cả",AVERAGE(TblData[Số lần mua]),AVERAGEIF(TblData[Phân khúc (k=4)],$E$5,TblData[Số lần mua]))', "0.0"),
        ("Số sản phẩm khác nhau TB",
         '=IF($E$5="Tất cả",AVERAGE(TblData[Số sản phẩm khác nhau]),AVERAGEIF(TblData[Phân khúc (k=4)],$E$5,TblData[Số sản phẩm khác nhau]))', "0.0"),
    ]
    kpi_row = 7
    for i, (label, formula, fmt) in enumerate(kpi_defs):
        r = kpi_row + i
        ws_dash[f"B{r}"] = label
        ws_dash[f"B{r}"].font = Font(name="Calibri", size=10.5, color="4A4A48")
        ws_dash[f"C{r}"] = formula
        ws_dash[f"C{r}"].font = Font(name="Calibri", bold=True, size=13, color=HEADER_FILL)
        ws_dash[f"C{r}"].number_format = fmt
        ws_dash[f"B{r}"].fill = PatternFill("solid", fgColor=LIGHT_FILL)
        ws_dash[f"C{r}"].fill = PatternFill("solid", fgColor=LIGHT_FILL)

    # Bảng phân bố theo cụm (cố định, không phụ thuộc dropdown) — nguồn cho pie chart
    dist_row = 14
    ws_dash[f"B{dist_row}"] = "Phân bố khách hàng theo phân khúc (k=4)"
    ws_dash[f"B{dist_row}"].font = Font(name="Calibri", bold=True, size=12)
    header_r = dist_row + 1
    ws_dash[f"B{header_r}"] = "Phân khúc"
    ws_dash[f"C{header_r}"] = "Số khách hàng"
    _style_header_row(ws_dash, header_r, 2)
    ws_dash.cell(row=header_r, column=2).alignment = Alignment(horizontal="left")
    for i, name in enumerate(persona_names):
        r = header_r + 1 + i
        ws_dash[f"B{r}"] = name
        ws_dash[f"C{r}"] = f'=COUNTIF(TblData[Phân khúc (k=4)],B{r})'

    # Bảng so sánh chỉ số theo cụm (cố định) — nguồn cho bar chart
    comp_row = header_r + 1 + len(persona_names) + 2
    ws_dash[f"B{comp_row}"] = "So sánh chỉ số trung bình theo phân khúc (k=4)"
    ws_dash[f"B{comp_row}"].font = Font(name="Calibri", bold=True, size=12)
    comp_header_r = comp_row + 1
    ws_dash[f"B{comp_header_r}"] = "Phân khúc"
    ws_dash[f"C{comp_header_r}"] = "Chi tiêu TB (£)"
    ws_dash[f"D{comp_header_r}"] = "Số lần mua TB"
    _style_header_row(ws_dash, comp_header_r, 3)
    ws_dash.cell(row=comp_header_r, column=2).alignment = Alignment(horizontal="left")
    for i, name in enumerate(persona_names):
        r = comp_header_r + 1 + i
        ws_dash[f"B{r}"] = name
        ws_dash[f"C{r}"] = f'=AVERAGEIF(TblData[Phân khúc (k=4)],B{r},TblData[Tổng chi tiêu])'
        ws_dash[f"D{r}"] = f'=AVERAGEIF(TblData[Phân khúc (k=4)],B{r},TblData[Số lần mua])'

    for col in ["B", "C", "D", "E"]:
        ws_dash.column_dimensions[col].width = 26

    # Pie chart — phân bố theo cụm
    pie = PieChart()
    pie.title = "Phân bố khách hàng theo phân khúc"
    data_ref = Reference(ws_dash, min_col=3, min_row=header_r, max_row=header_r + len(persona_names))
    cats_ref = Reference(ws_dash, min_col=2, min_row=header_r + 1, max_row=header_r + len(persona_names))
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(cats_ref)
    pie.height, pie.width = 9, 13
    ws_dash.add_chart(pie, "F7")

    # Bar chart — so sánh chỉ số theo cụm
    bar = BarChart()
    bar.type = "col"
    bar.title = "So sánh chi tiêu & tần suất trung bình theo phân khúc"
    bar.y_axis.title = None
    data_ref2 = Reference(ws_dash, min_col=3, max_col=4, min_row=comp_header_r,
                           max_row=comp_header_r + len(persona_names))
    cats_ref2 = Reference(ws_dash, min_col=2, min_row=comp_header_r + 1, max_row=comp_header_r + len(persona_names))
    bar.add_data(data_ref2, titles_from_data=True)
    bar.set_categories(cats_ref2)
    bar.height, bar.width = 9, 13
    ws_dash.add_chart(bar, "F25")

    # ══ Sheet "So Sánh Cụm" (16 đặc trưng, color scale) ══
    ws_cmp = wb.create_sheet("So Sánh Cụm")
    merged4 = features_df.copy()
    merged4.index = merged4.index.astype(str)
    if clusters_k4 is not None and len(clusters_k4) > 0:
        ck4 = clusters_k4.copy()
        ck4["CustomerID"] = ck4["CustomerID"].astype(str)
        merged4 = merged4.merge(ck4.set_index("CustomerID")["Cluster"], left_index=True, right_index=True, how="left")
        cluster_avg = merged4.groupby("Cluster")[list(features_df.columns)].mean()
        cluster_avg.index = [SEGMENT_PERSONAS_K4.get(int(c), {}).get("name", f"Cụm {c}") for c in cluster_avg.index]

        ws_cmp["A1"] = "Chỉ số"
        for j, cname in enumerate(cluster_avg.index):
            ws_cmp.cell(row=1, column=2 + j, value=cname)
        for i, feat in enumerate(features_df.columns):
            ws_cmp.cell(row=2 + i, column=1, value=FEATURE_NAMES_VN.get(feat, feat))
            for j, cname in enumerate(cluster_avg.index):
                ws_cmp.cell(row=2 + i, column=2 + j, value=round(float(cluster_avg.loc[cname, feat]), 2))

        n_feat, n_clust = len(features_df.columns), len(cluster_avg.index)
        _style_header_row(ws_cmp, 1, 1 + n_clust)
        ws_cmp.column_dimensions["A"].width = 28
        for j in range(n_clust):
            ws_cmp.column_dimensions[get_column_letter(2 + j)].width = 20

        for i in range(n_feat):
            r = 2 + i
            rule = ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B",
            )
            ws_cmp.conditional_formatting.add(
                f"B{r}:{get_column_letter(1 + n_clust)}{r}", rule
            )
    else:
        ws_cmp["A1"] = "Không có dữ liệu cụm k=4 để so sánh."

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────
# DASHBOARD IMAGE BUILDER (PNG/JPEG — ghép nhiều widget trên 1 canvas)
# ─────────────────────────────────────────────
def build_dashboard_image(cleaned_df, features_df, clusters_k3, clusters_k4, data_label, img_format="png"):
    """Ghép các 'widget' (KPI card, pie chart, bar chart, radar chart) thành MỘT ảnh dashboard
    duy nhất bằng matplotlib GridSpec, xuất PNG hoặc JPEG — không cần chụp màn hình trình duyệt."""
    from matplotlib import gridspec
    from matplotlib.patches import FancyBboxPatch

    n_customers = len(features_df)
    n_transactions = len(cleaned_df) if cleaned_df is not None else None
    total_revenue = features_df["Sum_TotalPrice"].sum() if "Sum_TotalPrice" in features_df.columns else None
    avg_spend = features_df["Sum_TotalPrice"].mean() if "Sum_TotalPrice" in features_df.columns else None
    avg_txn = features_df["Count_Invoice"].mean() if "Count_Invoice" in features_df.columns else None

    kpis = [("Khách hàng", f"{n_customers:,}")]
    if n_transactions is not None:
        kpis.append(("Giao dịch", f"{n_transactions:,}"))
    if total_revenue is not None:
        kpis.append(("Tổng doanh thu", f"£{total_revenue:,.0f}"))
    if avg_spend is not None:
        kpis.append(("Chi tiêu TB/khách", f"£{avg_spend:,.0f}"))
    if avg_txn is not None:
        kpis.append(("Giao dịch TB/khách", f"{avg_txn:,.1f}"))
    n_kpi = max(len(kpis), 1)

    NAVY, MUTED, BG = "#1E2761", "#6B7A99", "#F7F8FC"

    fig = plt.figure(figsize=(16, 11), dpi=160)
    fig.patch.set_facecolor(BG)
    gs = gridspec.GridSpec(
        4, 4, figure=fig, height_ratios=[0.5, 0.8, 2.3, 2.6],
        hspace=0.6, wspace=0.35, left=0.04, right=0.97, top=0.95, bottom=0.04,
    )

    # ── Header ──
    ax_h = fig.add_subplot(gs[0, :])
    ax_h.axis("off")
    ax_h.text(0, 0.75, "DASHBOARD PHÂN KHÚC KHÁCH HÀNG", fontsize=23, fontweight="bold", color=NAVY, va="top")
    ax_h.text(0, 0.15, f"Ngày xuất: {datetime.now().strftime('%d/%m/%Y %H:%M')}  ·  Nguồn dữ liệu: {data_label}",
              fontsize=11.5, color=MUTED, va="top")

    # ── KPI cards ──
    gs_kpi = gridspec.GridSpecFromSubplotSpec(1, n_kpi, subplot_spec=gs[1, :], wspace=0.25)
    for i, (label, value) in enumerate(kpis):
        ax = fig.add_subplot(gs_kpi[0, i])
        ax.axis("off")
        ax.set_xlim(0, 1)
        ax.set_ylim(0, 1)
        ax.add_patch(FancyBboxPatch(
            (0.02, 0.05), 0.96, 0.9, boxstyle="round,pad=0.02,rounding_size=0.06",
            transform=ax.transAxes, facecolor="white", edgecolor="#E2E6F2", linewidth=1.2,
        ))
        ax.text(0.5, 0.6, value, fontsize=21, fontweight="bold", color=NAVY, ha="center", va="center", transform=ax.transAxes)
        ax.text(0.5, 0.24, label, fontsize=11.5, color=MUTED, ha="center", va="center", transform=ax.transAxes)

    # ── Pie chart phân bố k=4 (trái) & Bar chart so sánh (phải) ──
    if clusters_k4 is not None and len(clusters_k4) > 0:
        counts = clusters_k4["Cluster"].value_counts().sort_index()
        total = counts.sum()
        labels = [SEGMENT_PERSONAS_K4.get(int(c), {}).get("name", f"Cụm {c}") for c in counts.index]
        colors_list = [SEGMENT_PERSONAS_K4.get(int(c), {}).get("color", "#888888") for c in counts.index]

        ax_pie = fig.add_subplot(gs[2, 0:2])
        wedges, _ = ax_pie.pie(counts.values, colors=colors_list, startangle=90,
                                wedgeprops=dict(width=0.42, edgecolor="white"))
        ax_pie.legend(wedges, [f"{l} ({c/total:.1%})" for l, c in zip(labels, counts.values)],
                      loc="center left", bbox_to_anchor=(1.0, 0.5), fontsize=9.5, frameon=False)
        ax_pie.set_title("Phân bố khách hàng theo phân khúc (k=4)", fontsize=13, color=NAVY, pad=12)

        ax_bar = fig.add_subplot(gs[2, 2:4])
        merged4 = features_df.copy()
        merged4.index = merged4.index.astype(str)
        ck4 = clusters_k4.copy()
        ck4["CustomerID"] = ck4["CustomerID"].astype(str)
        merged4 = merged4.merge(ck4.set_index("CustomerID")["Cluster"], left_index=True, right_index=True, how="left")
        avg_spend_c = merged4.groupby("Cluster")["Sum_TotalPrice"].mean() if "Sum_TotalPrice" in merged4.columns else None
        if avg_spend_c is not None:
            bar_labels = [SEGMENT_PERSONAS_K4.get(int(c), {}).get("name", f"Cụm {c}") for c in avg_spend_c.index]
            bar_colors = [SEGMENT_PERSONAS_K4.get(int(c), {}).get("color", "#888888") for c in avg_spend_c.index]
            ax_bar.bar(bar_labels, avg_spend_c.values, color=bar_colors)
            ax_bar.set_title("Chi tiêu trung bình theo phân khúc (£)", fontsize=13, color=NAVY, pad=12)
            ax_bar.tick_params(axis="x", labelrotation=12, labelsize=9)
            ax_bar.spines[["top", "right"]].set_visible(False)

    # ── Radar chart đặc trưng (dưới, căn giữa) ──
    if clusters_k4 is not None and len(clusters_k4) > 0:
        radar_keys = [f for f in RADAR_FEATURES.keys() if f in merged4.columns]
        cluster_means = merged4.groupby("Cluster")[radar_keys].mean()
        gmin, gmax = cluster_means.min(), cluster_means.max()
        norm = (cluster_means - gmin) / (gmax - gmin + 1e-9)
        categories = [RADAR_FEATURES[f] for f in radar_keys]
        angles = np.linspace(0, 2 * np.pi, len(categories), endpoint=False).tolist()
        angles += angles[:1]

        ax_radar = fig.add_subplot(gs[3, 1:3], projection="polar")
        for cid, row in norm.iterrows():
            vals = row.tolist() + row.tolist()[:1]
            color = SEGMENT_PERSONAS_K4.get(int(cid), {}).get("color", "#888888")
            name = SEGMENT_PERSONAS_K4.get(int(cid), {}).get("name", f"Cụm {cid}")
            ax_radar.plot(angles, vals, color=color, linewidth=1.8, label=name)
            ax_radar.fill(angles, vals, color=color, alpha=0.12)
        ax_radar.set_xticks(angles[:-1])
        ax_radar.set_xticklabels(categories, fontsize=9)
        ax_radar.set_yticklabels([])
        ax_radar.legend(loc="upper right", bbox_to_anchor=(1.45, 1.12), fontsize=9.5, frameon=False)
        ax_radar.set_title("So sánh đặc trưng trung bình từng phân khúc (đã chuẩn hóa)", fontsize=13, color=NAVY, pad=22)

    buf = io.BytesIO()
    save_format = "jpeg" if img_format.lower() in ("jpg", "jpeg") else "png"
    fig.savefig(buf, format=save_format, dpi=160, facecolor=fig.get_facecolor())
    plt.close(fig)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────
# PLOTLY THEME
# ─────────────────────────────────────────────
PLOTLY_LAYOUT = dict(
    paper_bgcolor="rgba(0,0,0,0)",
    plot_bgcolor="rgba(0,0,0,0)",
    font=dict(color="#edf0f5", family="DM Sans"),
    xaxis=dict(gridcolor="#262c3a", zerolinecolor="#262c3a"),
    yaxis=dict(gridcolor="#262c3a", zerolinecolor="#262c3a"),
    margin=dict(l=40, r=20, t=40, b=40),
)


def render_page_hero(kicker, title, subtitle, chips):
        chip_html = "".join(f"<span class='hero-chip'>{chip}</span>" for chip in chips)
        st.markdown(
                f"""
<div class='hero-banner'>
    <div class='hero-kicker'>{kicker}</div>
    <div class='hero-title'>{title}</div>
    <div class='hero-subtitle'>{subtitle}</div>
    <div class='hero-chips'>{chip_html}</div>
</div>
""",
                unsafe_allow_html=True,
        )


# ─────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 🎯 Customer Segmentation")
    st.markdown("<div class='section-header'>Cài đặt phân tích</div>", unsafe_allow_html=True)

    k_choice = st.radio("Số lượng phân khúc", options=[3, 4], index=1, horizontal=True)
    st.markdown("---")

    st.markdown("<div class='section-header'>Navigation</div>", unsafe_allow_html=True)
    page = st.radio(
        "Chọn trang",
        ["📊 Tổng Quan", "🔍 Khám Phá Phân Khúc", "⚖️ So Sánh K=3 vs K=4", "📈 Phân Tích RFM", "🧭 Không Gian PCA", "🏆 Xếp Hạng Khách Hàng", "👤 Tra Cứu Khách Hàng", "🧪 Mô Phỏng Khách Hàng", "📤 Cập Nhật Dữ Liệu Mới", "📄 Xuất Báo Cáo Tổng Hợp"],
        label_visibility="collapsed",
    )
    st.markdown("---")

    if st.session_state.get("recomputed"):
        rec = st.session_state["recomputed"]
        new_label = f"+{rec['n_new_customers']} khách hàng mới · " if rec.get("n_new_customers") is not None else ""
        st.markdown(
            f"<div style='color:#2dd4a7;font-size:0.75rem;'>🟢 Đang dùng dữ liệu đã gộp<br>"
            f"{new_label}{rec['n_total_customers']} tổng cộng</div>",
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            "<div style='color:#8894a8;font-size:0.72rem;'>⚪ Đang dùng dữ liệu gốc</div>",
            unsafe_allow_html=True,
        )
    st.markdown(
        "<div style='color:#8894a8;font-size:0.72rem;margin-top:8px;'>Online Retail Dataset · UK Customers<br>KMeans + PCA + Feature Engineering</div>",
        unsafe_allow_html=True,
    )

# ─────────────────────────────────────────────
# LOAD DATA  (dùng dữ liệu đã gộp/tính lại nếu có)
# ─────────────────────────────────────────────
data = load_all_data()

# Nếu trang vừa được tải lại (F5) và session_state trống, tự động khôi phục
# dữ liệu đã gộp lần gần nhất từ file *_merged.csv trên đĩa (nếu có).
if "recomputed" not in st.session_state:
    restored = load_merged_from_disk()
    st.session_state["recomputed"] = restored
    st.session_state["just_restored"] = restored is not None

if st.session_state.get("recomputed"):
    rec = st.session_state["recomputed"]
    data["cleaned"] = rec["cleaned"]
    data["features"] = rec["features"]
    data["scaled"] = rec["scaled"]
    data["clusters_k3"] = rec["clusters_k3"]
    data["clusters_k4"] = rec["clusters_k4"]

features_df = data["features"]
scaled_df = data["scaled"]
clusters_k = data[f"clusters_k{k_choice}"]
merged_df = merge_clusters(features_df, clusters_k, k_choice)
cleaned_df = data["cleaned"]
colors = get_cluster_colors(k_choice)
personas = get_personas(k_choice)
cluster_col = f"Cluster_{k_choice}"

# Check data availability
if features_df is None or clusters_k is None:
    st.error("⚠️ Không tìm thấy dữ liệu. Hãy chạy pipeline và đảm bảo các file CSV tồn tại trong `data/processed/`.")
    st.info(
        "Các file cần thiết:\n"
        "- `data/processed/customer_features.csv`\n"
        "- `data/processed/customer_clusters_k3.csv`\n"
        "- `data/processed/customer_clusters_k4.csv`\n"
        "- `data/processed/cleaned_uk_data.csv` (tùy chọn)"
    )
    st.stop()


# ─────────────────────────────────────────────
# ══ PAGE: TỔNG QUAN ══
# ─────────────────────────────────────────────
if page == "📊 Tổng Quan":
    st.markdown("# 📊 Tổng Quan")
    render_page_hero(
        "Customer Segmentation Dashboard",
        "Phân tích hành vi khách hàng bằng K-Means, PCA và SHAP",
        "Giao diện tập trung vào 16 features cấp khách hàng, giúp xem nhanh quy mô từng nhóm, so sánh radar chart và truy vết nguyên nhân bằng SHAP.",
        ["K-Means clustering", "PCA visualization", "16 SHAP features", "Business personas"],
    )
    st.markdown(f"<div class='section-header'>Phân tích K = {k_choice} phân khúc khách hàng</div>", unsafe_allow_html=True)

    # ── Top KPI row
    if merged_df is not None and cluster_col in merged_df.columns:
        total_customers = len(merged_df)
        cluster_sizes = merged_df[cluster_col].value_counts().sort_index()
        largest_cluster_pct = (cluster_sizes.max() / total_customers * 100)
        avg_spend = merged_df["Sum_TotalPrice"].mean() if "Sum_TotalPrice" in merged_df.columns else 0

        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Tổng khách hàng", f"{total_customers:,}")
        with col2:
            st.metric("Số phân khúc", k_choice)
        with col3:
            st.metric("Chi tiêu TB (GBP)", f"£{avg_spend:,.0f}")
        with col4:
            st.metric("Phân khúc lớn nhất", f"{largest_cluster_pct:.1f}%")

        st.markdown("---")

        # ── Persona cards + distribution
        st.markdown("### Phân Khúc Khách Hàng")
        for cid, persona in personas.items():
            count = cluster_sizes.get(cid, 0)
            pct = count / total_customers * 100
            col_card, col_bar = st.columns([3, 7])
            with col_card:
                st.markdown(
                    f"""<div class='metric-card' style='border-left: 4px solid {persona["color"]}'>
                        <div class='metric-value' style='color:{persona["color"]};font-size:1.4rem'>{persona["name"]}</div>
                        <div class='metric-label'>{persona["desc"]}</div>
                        <div style='margin-top:8px;font-family:Space Mono,monospace;font-size:1.1rem'>{count:,} KH · {pct:.1f}%</div>
                    </div>""",
                    unsafe_allow_html=True,
                )
            with col_bar:
                bar_fig = go.Figure(go.Bar(
                    x=[count], y=[persona["name"]], orientation="h",
                    marker_color=persona["color"], text=[f"{count:,} ({pct:.1f}%)"],
                    textposition="outside", textfont=dict(color="#edf0f5"),
                ))
                bar_fig.update_layout(
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    font=dict(color="#edf0f5", family="DM Sans"),
                    height=80, margin=dict(l=0, r=40, t=8, b=8),
                    xaxis=dict(showticklabels=False, showgrid=False, zeroline=False, zerolinecolor="#262c3a"),
                    yaxis=dict(showticklabels=False),
                )
                st.plotly_chart(bar_fig, use_container_width=True)

        st.markdown("---")

        # ── Cluster size pie + feature heatmap
        col_pie, col_heat = st.columns(2)
        with col_pie:
            st.markdown("#### Tỷ lệ phân khúc")
            pie_fig = go.Figure(go.Pie(
                labels=[personas[i]["name"] for i in cluster_sizes.index if i in personas],
                values=cluster_sizes.values,
                marker_colors=[personas[i]["color"] for i in cluster_sizes.index if i in personas],
                hole=0.5,
                textinfo="label+percent",
                textfont=dict(size=11),
            ))
            pie_fig.update_layout(**PLOTLY_LAYOUT, height=320, showlegend=False)
            st.plotly_chart(pie_fig, use_container_width=True)

        with col_heat:
            st.markdown("#### Feature nổi bật theo phân khúc")
            key_features = ["Sum_TotalPrice", "Count_Invoice", "Sum_Quantity", "Mean_UnitPrice", "Count_Stock"]
            available_kf = [f for f in key_features if f in merged_df.columns]
            cluster_means = merged_df.groupby(cluster_col)[available_kf].mean()
            norm = (cluster_means - cluster_means.min()) / (cluster_means.max() - cluster_means.min() + 1e-9)
            heat_fig = go.Figure(go.Heatmap(
                z=norm.values,
                x=[FEATURE_NAMES_VN.get(c, c) for c in norm.columns],
                y=[personas.get(int(i), {}).get("name", f"Cluster {i}") for i in norm.index],
                colorscale=[[0, "#171c28"], [0.5, "#7c8aff"], [1, "#ff8fa3"]],
                text=np.round(cluster_means.values, 1),
                texttemplate="%{text:.0f}",
                textfont=dict(size=10),
            ))
            heat_fig.update_layout(**PLOTLY_LAYOUT, height=320)
            st.plotly_chart(heat_fig, use_container_width=True)


# ─────────────────────────────────────────────
# ══ PAGE: KHÁM PHÁ PHÂN KHÚC ══
# ─────────────────────────────────────────────
elif page == "🔍 Khám Phá Phân Khúc":
    st.markdown("# 🔍 Khám Phá Phân Khúc")
    render_page_hero(
        "Segmentation Explorer",
        "So sánh radar chart, feature distribution và bảng thống kê",
        "Mục này được thiết kế như một bảng điều khiển khám phá: xem nhanh profile từng cluster rồi drill-down theo feature cụ thể.",
        ["Radar chart", "Box / violin plots", "Cluster statistics"],
    )

    if merged_df is None or cluster_col not in merged_df.columns:
        st.warning("Không có dữ liệu cluster.")
        st.stop()

    tab1, tab2, tab3 = st.tabs(["📡 Radar Chart", "📊 So Sánh Feature", "📋 Bảng Thống Kê"])

    # ── TAB 1: Radar Chart
    with tab1:
        radar_keys = [f for f in RADAR_FEATURES.keys() if f in merged_df.columns]
        cluster_means = merged_df.groupby(cluster_col)[radar_keys].mean()
        global_min = cluster_means.min()
        global_max = cluster_means.max()
        norm = (cluster_means - global_min) / (global_max - global_min + 1e-9)

        categories = [RADAR_FEATURES[f] for f in radar_keys]
        fig2 = go.Figure()
        for cid, row in norm.iterrows():
            vals = row.tolist() + row.tolist()[:1]
            cats = categories + categories[:1]
            color = colors.get(int(cid), "#ffffff")
            pname = personas.get(int(cid), {}).get("name", f"Cluster {cid}")
            h = color.lstrip("#")
            r2, g2, b2 = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
            fig2.add_trace(go.Scatterpolar(
                r=vals, theta=cats, name=pname,
                fill="toself",
                fillcolor=f"rgba({r2},{g2},{b2},0.15)",
                line=dict(color=color, width=3),
                marker=dict(size=8, color=color),
            ))

        fig2.update_layout(
            **PLOTLY_LAYOUT,
            height=480,
            polar=dict(
                bgcolor="rgba(0,0,0,0)",
                radialaxis=dict(visible=True, range=[0, 1], gridcolor="#262c3a", color="#8894a8"),
                angularaxis=dict(gridcolor="#262c3a", color="#edf0f5"),
            ),
            legend=dict(orientation="h", y=-0.12, x=0.5, xanchor="center"),
        )
        st.plotly_chart(fig2, use_container_width=True)

    # ── TAB 2: Feature comparison
    with tab2:
        feature_options = {FEATURE_NAMES_VN.get(f, f): f for f in features_df.columns if f in merged_df.columns}
        selected_label = st.selectbox("Chọn feature để so sánh", list(feature_options.keys()), index=3)
        selected_feat = feature_options[selected_label]

        col_box, col_violin = st.columns(2)
        with col_box:
            fig_box = go.Figure()
            for cid in sorted(merged_df[cluster_col].dropna().unique()):
                cid_int = int(cid)
                subset = merged_df[merged_df[cluster_col] == cid][selected_feat].dropna()
                p99 = subset.quantile(0.99)
                subset = subset[subset <= p99]
                fig_box.add_trace(go.Box(
                    y=subset, name=personas.get(cid_int, {}).get("name", f"C{cid}"),
                    marker_color=colors.get(cid_int, "#fff"), boxmean=True,
                ))
            fig_box.update_layout(**PLOTLY_LAYOUT, height=360, title="Box Plot")
            st.plotly_chart(fig_box, use_container_width=True)

        with col_violin:
            fig_vio = go.Figure()
            for cid in sorted(merged_df[cluster_col].dropna().unique()):
                cid_int = int(cid)
                subset = merged_df[merged_df[cluster_col] == cid][selected_feat].dropna()
                p99 = subset.quantile(0.99)
                subset = subset[subset <= p99]
                fig_vio.add_trace(go.Violin(
                    y=subset, name=personas.get(cid_int, {}).get("name", f"C{cid}"),
                    fillcolor=colors.get(cid_int, "#fff"),
                    line_color=colors.get(cid_int, "#fff"),
                    opacity=0.7, box_visible=True, meanline_visible=True,
                ))
            fig_vio.update_layout(**PLOTLY_LAYOUT, height=360, title="Violin Plot")
            st.plotly_chart(fig_vio, use_container_width=True)

        # Multi-feature bar comparison
        st.markdown("#### So sánh nhiều feature giữa các phân khúc")
        multi_feat_labels = st.multiselect(
            "Chọn các feature",
            list(feature_options.keys()),
            default=list(feature_options.keys())[:5],
        )
        if multi_feat_labels:
            multi_feats = [feature_options[l] for l in multi_feat_labels]
            avail = [f for f in multi_feats if f in merged_df.columns]
            cm = merged_df.groupby(cluster_col)[avail].mean()
            norm_cm = (cm - cm.min()) / (cm.max() - cm.min() + 1e-9)

            fig_bar = go.Figure()
            for cid in sorted(norm_cm.index):
                cid_int = int(cid)
                fig_bar.add_trace(go.Bar(
                    name=personas.get(cid_int, {}).get("name", f"C{cid}"),
                    x=[FEATURE_NAMES_VN.get(f, f) for f in avail],
                    y=norm_cm.loc[cid].values,
                    marker_color=colors.get(cid_int, "#fff"),
                    opacity=0.85,
                ))
            fig_bar.update_layout(**PLOTLY_LAYOUT, height=380, barmode="group",
                                  xaxis_tickangle=-30, legend=dict(orientation="h", y=-0.2))
            st.plotly_chart(fig_bar, use_container_width=True)

    # ── TAB 3: Stats table
    with tab3:
        st.markdown("#### Thống kê mô tả theo phân khúc")
        disp_feats = [f for f in features_df.columns if f in merged_df.columns]
        stats = merged_df.groupby(cluster_col)[disp_feats].agg(["mean", "median", "std"]).round(2)
        stats.columns = [f"{FEATURE_NAMES_VN.get(c, c)} ({s})" for c, s in stats.columns]
        stats.index = [personas.get(int(i), {}).get("name", f"Cluster {i}") for i in stats.index]
        st.dataframe(stats.T, use_container_width=True, height=500)


# ─────────────────────────────────────────────
# ══ PAGE: SO SÁNH K=3 vs K=4 ══
# ─────────────────────────────────────────────
elif page == "⚖️ So Sánh K=3 vs K=4":
    st.markdown("# ⚖️ So Sánh K=3 vs K=4")
    render_page_hero(
        "Algorithm Comparison",
        "K=3 tách thành K=4 như thế nào?",
        "Trang này trực quan hóa việc tăng số phân khúc từ 3 lên 4 đã tách nhỏ nhóm nào, giúp lý giải lựa chọn K=4 làm kết quả chính thức của luận văn.",
        ["Sankey flow", "So sánh song song", "Segment migration"],
    )

    merged_k3 = merge_clusters(features_df, data.get("clusters_k3"), 3)
    merged_k4 = merge_clusters(features_df, data.get("clusters_k4"), 4)

    if merged_k3 is None or merged_k4 is None:
        st.warning("Cần cả dữ liệu K=3 và K=4 để so sánh. Hãy đảm bảo cả hai file CSV đều tồn tại.")
        st.stop()

    combo = merged_k3[["Cluster_3"]].join(merged_k4[["Cluster_4"]], how="inner").dropna()
    combo["Cluster_3"] = combo["Cluster_3"].astype(int)
    combo["Cluster_4"] = combo["Cluster_4"].astype(int)

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Số khách hàng chung", f"{len(combo):,}")
    with col2:
        st.metric("Số phân khúc K=3", 3)
    with col3:
        st.metric("Số phân khúc K=4", 4)

    st.markdown("---")
    st.markdown("### Luồng di chuyển khách hàng: K=3 → K=4")

    k3_ids = sorted(SEGMENT_PERSONAS_K3.keys())
    k4_ids = sorted(SEGMENT_PERSONAS_K4.keys())
    labels = [f"K3 · {SEGMENT_PERSONAS_K3[i]['name']}" for i in k3_ids] + \
             [f"K4 · {SEGMENT_PERSONAS_K4[i]['name']}" for i in k4_ids]
    node_colors = [SEGMENT_PERSONAS_K3[i]["color"] for i in k3_ids] + \
                  [SEGMENT_PERSONAS_K4[i]["color"] for i in k4_ids]

    sources, targets, values, link_colors = [], [], [], []
    for k3_id in k3_ids:
        for k4_id in k4_ids:
            count = int(((combo["Cluster_3"] == k3_id) & (combo["Cluster_4"] == k4_id)).sum())
            if count > 0:
                sources.append(k3_ids.index(k3_id))
                targets.append(len(k3_ids) + k4_ids.index(k4_id))
                values.append(count)
                h = SEGMENT_PERSONAS_K4[k4_id]["color"].lstrip("#")
                r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
                link_colors.append(f"rgba({r},{g},{b},0.45)")

    fig_sankey = go.Figure(go.Sankey(
        node=dict(label=labels, color=node_colors, pad=20, thickness=22,
                  line=dict(color="#262c3a", width=1)),
        link=dict(source=sources, target=targets, value=values, color=link_colors),
    ))
    fig_sankey.update_layout(
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        font=dict(color="#edf0f5", family="DM Sans", size=12),
        height=460, margin=dict(l=10, r=10, t=20, b=10),
    )
    st.plotly_chart(fig_sankey, use_container_width=True)
    st.caption("Độ rộng luồng thể hiện số khách hàng di chuyển từ phân khúc K=3 sang phân khúc K=4 tương ứng.")

    st.markdown("---")
    col_k3, col_k4 = st.columns(2)
    with col_k3:
        st.markdown("#### Phân bố K=3")
        sizes3 = combo["Cluster_3"].value_counts().sort_index()
        fig_p3 = go.Figure(go.Pie(
            labels=[SEGMENT_PERSONAS_K3[i]["name"] for i in sizes3.index],
            values=sizes3.values,
            marker_colors=[SEGMENT_PERSONAS_K3[i]["color"] for i in sizes3.index],
            hole=0.5, textinfo="label+percent", textfont=dict(size=10),
        ))
        fig_p3.update_layout(**PLOTLY_LAYOUT, height=320, showlegend=False)
        st.plotly_chart(fig_p3, use_container_width=True)
    with col_k4:
        st.markdown("#### Phân bố K=4")
        sizes4 = combo["Cluster_4"].value_counts().sort_index()
        fig_p4 = go.Figure(go.Pie(
            labels=[SEGMENT_PERSONAS_K4[i]["name"] for i in sizes4.index],
            values=sizes4.values,
            marker_colors=[SEGMENT_PERSONAS_K4[i]["color"] for i in sizes4.index],
            hole=0.5, textinfo="label+percent", textfont=dict(size=10),
        ))
        fig_p4.update_layout(**PLOTLY_LAYOUT, height=320, showlegend=False)
        st.plotly_chart(fig_p4, use_container_width=True)

    st.markdown("---")
    st.markdown("### Nhận xét")
    crosstab = pd.crosstab(combo["Cluster_3"], combo["Cluster_4"])
    any_insight = False
    for k3_id in crosstab.index:
        row = crosstab.loc[k3_id]
        row_nonzero = row[row > 0]
        if len(row_nonzero) > 1:
            any_insight = True
            k3_name = SEGMENT_PERSONAS_K3[int(k3_id)]["name"]
            splits_desc = ", ".join(
                f"**{SEGMENT_PERSONAS_K4[int(k4_id)]['name']}** ({cnt:,} KH)"
                for k4_id, cnt in row_nonzero.items()
            )
            st.markdown(f"- Nhóm **{k3_name}** (K=3) được tách thành: {splits_desc}")
    if not any_insight:
        st.info("Mỗi phân khúc K=3 ánh xạ gần như trọn vẹn sang một phân khúc K=4 tương ứng, không có sự tách nhóm rõ rệt.")


# ─────────────────────────────────────────────
# ══ PAGE: PHÂN TÍCH RFM ══
# ─────────────────────────────────────────────
elif page == "📈 Phân Tích RFM":
    st.markdown("# 📈 Phân Tích RFM")
    render_page_hero(
        "Retail Activity",
        "Quan sát doanh thu, sản phẩm và hành vi mua theo thời gian",
        "Trang này giúp nhìn dữ liệu gốc theo góc vận hành: thời gian, sản phẩm bán chạy và phân phối giao dịch của khách hàng.",
        ["Revenue trend", "Top products", "Customer distribution"],
    )

    if cleaned_df is None:
        st.warning("Không tìm thấy `cleaned_uk_data.csv`. Trang này cần dữ liệu giao dịch thô.")
        st.stop()

    # Ensure datetime
    if "InvoiceDate" in cleaned_df.columns:
        cleaned_df["InvoiceDate"] = pd.to_datetime(cleaned_df["InvoiceDate"])

    tab_rfm1, tab_rfm2, tab_rfm3 = st.tabs(["📅 Doanh Thu Theo Thời Gian", "🛍️ Sản Phẩm", "🌍 Phân Phối Khách Hàng"])

    with tab_rfm1:
        col_d, col_m = st.columns(2)
        with col_d:
            st.markdown("#### Doanh thu hàng ngày")
            daily = cleaned_df.groupby(cleaned_df["InvoiceDate"].dt.date)["TotalPrice"].sum().reset_index()
            daily.columns = ["Date", "Revenue"]
            fig_daily = px.line(daily, x="Date", y="Revenue", color_discrete_sequence=["#7c8aff"])
            fig_daily.update_layout(**PLOTLY_LAYOUT, height=300, xaxis_title="Ngày", yaxis_title="Doanh thu (GBP)")
            st.plotly_chart(fig_daily, use_container_width=True)

        with col_m:
            st.markdown("#### Doanh thu hàng tháng")
            monthly = cleaned_df.groupby(cleaned_df["InvoiceDate"].dt.to_period("M").astype(str))["TotalPrice"].sum().reset_index()
            monthly.columns = ["Month", "Revenue"]
            fig_monthly = px.bar(monthly, x="Month", y="Revenue", color_discrete_sequence=["#ff8fa3"])
            fig_monthly.update_layout(**PLOTLY_LAYOUT, height=300, xaxis_tickangle=-45)
            st.plotly_chart(fig_monthly, use_container_width=True)

        st.markdown("#### Heatmap hoạt động theo ngày & giờ")
        if "DayOfWeek" in cleaned_df.columns and "HourOfDay" in cleaned_df.columns:
            heatmap_data = cleaned_df.groupby(["DayOfWeek", "HourOfDay"]).size().unstack(fill_value=0)
        else:
            heatmap_data = cleaned_df.copy()
            heatmap_data["DayOfWeek"] = heatmap_data["InvoiceDate"].dt.dayofweek
            heatmap_data["HourOfDay"] = heatmap_data["InvoiceDate"].dt.hour
            heatmap_data = heatmap_data.groupby(["DayOfWeek", "HourOfDay"]).size().unstack(fill_value=0)

        day_labels = ["Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6", "Thứ 7", "CN"]
        fig_heat = go.Figure(go.Heatmap(
            z=heatmap_data.values,
            x=[f"{h}:00" for h in heatmap_data.columns],
            y=[day_labels[d] if d < len(day_labels) else str(d) for d in heatmap_data.index],
            colorscale=[[0, "#171c28"], [0.5, "#7c8aff"], [1, "#ff8fa3"]],
        ))
        fig_heat.update_layout(**PLOTLY_LAYOUT, height=280, xaxis_title="Giờ trong ngày", yaxis_title="")
        st.plotly_chart(fig_heat, use_container_width=True)

    with tab_rfm2:
        top_n = st.slider("Số sản phẩm hiển thị", 5, 20, 10)
        col_qty, col_rev = st.columns(2)

        with col_qty:
            st.markdown(f"#### Top {top_n} sản phẩm theo số lượng")
            top_qty = cleaned_df.groupby("Description")["Quantity"].sum().sort_values(ascending=False).head(top_n)
            fig_qty = px.bar(
                x=top_qty.values, y=top_qty.index, orientation="h",
                color_discrete_sequence=["#7c8aff"],
                labels={"x": "Số lượng", "y": ""},
            )
            fig_qty.update_layout(**PLOTLY_LAYOUT, height=400)
            st.plotly_chart(fig_qty, use_container_width=True)

        with col_rev:
            st.markdown(f"#### Top {top_n} sản phẩm theo doanh thu")
            top_rev = cleaned_df.groupby("Description")["TotalPrice"].sum().sort_values(ascending=False).head(top_n)
            fig_rev = px.bar(
                x=top_rev.values, y=top_rev.index, orientation="h",
                color_discrete_sequence=["#ff8fa3"],
                labels={"x": "Doanh thu (GBP)", "y": ""},
            )
            fig_rev.update_layout(**PLOTLY_LAYOUT, height=400)
            st.plotly_chart(fig_rev, use_container_width=True)

    with tab_rfm3:
        st.markdown("#### Phân phối số giao dịch / khách hàng")
        txn_per_cust = cleaned_df.groupby("CustomerID")["InvoiceNo"].nunique()
        fig_hist = px.histogram(txn_per_cust[txn_per_cust <= txn_per_cust.quantile(0.99)],
                                nbins=40, color_discrete_sequence=["#5eb8ff"])
        fig_hist.update_layout(**PLOTLY_LAYOUT, height=300, xaxis_title="Số giao dịch", yaxis_title="Số khách hàng")
        st.plotly_chart(fig_hist, use_container_width=True)

        st.markdown("#### Phân phối tổng chi tiêu / khách hàng")
        spend_per_cust = cleaned_df.groupby("CustomerID")["TotalPrice"].sum()
        fig_spend = px.histogram(spend_per_cust[spend_per_cust <= spend_per_cust.quantile(0.99)],
                                 nbins=40, color_discrete_sequence=["#e8825a"])
        fig_spend.update_layout(**PLOTLY_LAYOUT, height=300, xaxis_title="Tổng chi tiêu (GBP)", yaxis_title="Số khách hàng")
        st.plotly_chart(fig_spend, use_container_width=True)


# ─────────────────────────────────────────────
# ══ PAGE: KHÔNG GIAN PCA ══
# ─────────────────────────────────────────────
elif page == "🧭 Không Gian PCA":
    st.markdown("# 🧭 Không Gian PCA")
    render_page_hero(
        "Dimensional View",
        "Khám phá cấu trúc cụm trong không gian PCA",
        "PCA giúp nhìn rõ độ tách cụm trong 2D/3D và cho biết bao nhiêu phương sai được giữ lại khi giảm chiều.",
        ["2D scatter", "3D scatter", "Explained variance"],
    )

    if scaled_df is None or merged_df is None:
        st.warning("Không có dữ liệu scaled features.")
        st.stop()

    feat_cols = [c for c in scaled_df.columns if not c.startswith("Cluster_")]
    X = scaled_df[feat_cols].values

    @st.cache_data
    def compute_pca(X_data, n=3):
        pca = PCA(n_components=n)
        comps = pca.fit_transform(X_data)
        return comps, pca.explained_variance_ratio_

    pca_comps, evr = compute_pca(X)
    pca_df = pd.DataFrame(pca_comps, columns=["PC1", "PC2", "PC3"], index=scaled_df.index)

    # Merge cluster labels
    if cluster_col in merged_df.columns:
        pca_df[cluster_col] = merged_df[cluster_col].values

    tab_2d, tab_3d, tab_var = st.tabs(["2D Scatter", "3D Scatter", "Phương Sai PCA"])

    with tab_2d:
        st.markdown(f"PC1 ({evr[0]:.1%}) · PC2 ({evr[1]:.1%}) — Tổng: {evr[0]+evr[1]:.1%}")
        fig_2d = go.Figure()
        for cid in sorted(pca_df[cluster_col].dropna().unique()):
            cid_int = int(cid)
            sub = pca_df[pca_df[cluster_col] == cid]
            fig_2d.add_trace(go.Scatter(
                x=sub["PC1"], y=sub["PC2"],
                mode="markers",
                name=personas.get(cid_int, {}).get("name", f"C{cid}"),
                marker=dict(color=colors.get(cid_int, "#fff"), size=5, opacity=0.7),
            ))
        fig_2d.update_layout(**PLOTLY_LAYOUT, height=480, xaxis_title="PC1", yaxis_title="PC2",
                              legend=dict(orientation="h", y=-0.12, x=0.5, xanchor="center"))
        st.plotly_chart(fig_2d, use_container_width=True)

    with tab_3d:
        st.markdown(f"PC1 · PC2 · PC3 — Tổng: {evr[:3].sum():.1%}")
        fig_3d = go.Figure()
        for cid in sorted(pca_df[cluster_col].dropna().unique()):
            cid_int = int(cid)
            sub = pca_df[pca_df[cluster_col] == cid]
            fig_3d.add_trace(go.Scatter3d(
                x=sub["PC1"], y=sub["PC2"], z=sub["PC3"],
                mode="markers",
                name=personas.get(cid_int, {}).get("name", f"C{cid}"),
                marker=dict(color=colors.get(cid_int, "#fff"), size=3, opacity=0.7),
            ))
        fig_3d.update_layout(
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)",
            font=dict(color="#edf0f5"),
            height=520,
            scene=dict(
                bgcolor="rgba(0,0,0,0)",
                xaxis=dict(title="PC1", gridcolor="#262c3a", color="#8894a8"),
                yaxis=dict(title="PC2", gridcolor="#262c3a", color="#8894a8"),
                zaxis=dict(title="PC3", gridcolor="#262c3a", color="#8894a8"),
            ),
            legend=dict(orientation="h", y=-0.05),
        )
        st.plotly_chart(fig_3d, use_container_width=True)

    with tab_var:
        full_pca_comps, full_evr = compute_pca(X, n=min(16, X.shape[1]))
        cum_evr = np.cumsum(full_evr)
        fig_var = go.Figure()
        fig_var.add_trace(go.Bar(
            x=list(range(1, len(full_evr) + 1)), y=full_evr,
            name="Phương sai riêng lẻ", marker_color="#7c8aff", opacity=0.8,
        ))
        fig_var.add_trace(go.Scatter(
            x=list(range(1, len(cum_evr) + 1)), y=cum_evr,
            name="Phương sai tích lũy", line=dict(color="#ff8fa3", width=2), mode="lines+markers",
        ))
        fig_var.add_hline(y=0.8, line_dash="dash", line_color="#5eb8ff", annotation_text="80%")
        fig_var.add_hline(y=0.9, line_dash="dash", line_color="#e8825a", annotation_text="90%")
        fig_var.update_layout(**PLOTLY_LAYOUT, height=380,
                              xaxis_title="Thành phần chính", yaxis_title="Tỷ lệ phương sai",
                              legend=dict(orientation="h", y=-0.15))
        st.plotly_chart(fig_var, use_container_width=True)

        pc_table = pd.DataFrame({
            "Thành phần": [f"PC{i+1}" for i in range(len(full_evr))],
            "Phương sai (%)": [f"{v:.2%}" for v in full_evr],
            "Tích lũy (%)": [f"{v:.2%}" for v in cum_evr],
        })
        st.dataframe(pc_table, use_container_width=True, hide_index=True)


# ─────────────────────────────────────────────
# ══ PAGE: XẾP HẠNG KHÁCH HÀNG ══
# ─────────────────────────────────────────────
elif page == "🏆 Xếp Hạng Khách Hàng":
    st.markdown("# 🏆 Xếp Hạng Khách Hàng")
    st.markdown("<div class='section-header'>Danh sách chi tiết khách hàng theo từng phân khúc</div>", unsafe_allow_html=True)
    render_page_hero(
        "Customer Ranking",
        "Xếp hạng theo hành vi mua và drill-down từng phân khúc",
        "Trang này phù hợp khi cần lọc, so sánh và xuất danh sách khách hàng theo các KPI kinh doanh.",
        ["Top customers", "Segment drill-down", "CSV export"],
    )

    if merged_df is None or cluster_col not in merged_df.columns:
        st.warning("Không có dữ liệu cluster.")
        st.stop()

    tab_rank, tab_segment = st.tabs(["🥇 Xếp Hạng Chi Tiêu", "📋 Danh Sách Theo Phân Khúc"])

    # ── TAB 1: Xếp hạng chi tiêu ──────────────────────────────────
    with tab_rank:
        rank_df = merged_df.copy().reset_index()
        rank_df = rank_df.rename(columns={"index": "CustomerID"})

        # Gắn tên cluster
        rank_df["Phân Khúc"] = rank_df[cluster_col].apply(
            lambda x: personas.get(int(x), {}).get("name", f"Cluster {x}") if pd.notna(x) else "N/A"
        )
        rank_df["Màu"] = rank_df[cluster_col].apply(
            lambda x: colors.get(int(x), "#8894a8") if pd.notna(x) else "#8894a8"
        )

        # Controls
        col_ctrl1, col_ctrl2, col_ctrl3 = st.columns([3, 3, 2])
        with col_ctrl1:
            sort_feature_label = st.selectbox(
                "Xếp hạng theo",
                ["Tổng chi tiêu", "Số lần mua", "Tổng số lượng", "Giá trung bình", "Số sản phẩm khác nhau"],
            )
        with col_ctrl2:
            filter_cluster = st.multiselect(
                "Lọc phân khúc",
                options=[personas.get(i, {}).get("name", f"C{i}") for i in sorted(merged_df[cluster_col].dropna().unique().astype(int))],
                default=[personas.get(i, {}).get("name", f"C{i}") for i in sorted(merged_df[cluster_col].dropna().unique().astype(int))],
            )
        with col_ctrl3:
            sort_order = st.radio("Thứ tự", ["Cao → Thấp", "Thấp → Cao"], horizontal=True)
            top_n_rank = st.slider("Hiển thị Top", min_value=10, max_value=500, value=50, step=10)

        # Map label → feature column
        sort_map = {
            "Tổng chi tiêu": "Sum_TotalPrice",
            "Số lần mua": "Count_Invoice",
            "Tổng số lượng": "Sum_Quantity",
            "Giá trung bình": "Mean_UnitPrice",
            "Số sản phẩm khác nhau": "Count_Stock",
        }
        sort_col = sort_map[sort_feature_label]

        # Filter + sort
        ascending = sort_order == "Thấp → Cao"
        filtered = rank_df[rank_df["Phân Khúc"].isin(filter_cluster)]
        if sort_col in filtered.columns:
            filtered = filtered.sort_values(sort_col, ascending=ascending)

        filtered = filtered.head(top_n_rank).reset_index(drop=True)
        filtered.index = filtered.index + 1  # bắt đầu từ 1

        # ── Bar chart xếp hạng
        st.markdown(f"#### Top {top_n_rank} khách hàng theo {sort_feature_label}")
        if sort_col in filtered.columns and len(filtered) > 0:
            fig_rank = go.Figure()
            for cid_int in sorted(merged_df[cluster_col].dropna().unique().astype(int)):
                sub = filtered[filtered[cluster_col] == cid_int]
                if len(sub) == 0:
                    continue
                pname = personas.get(cid_int, {}).get("name", f"C{cid_int}")
                color = colors.get(cid_int, "#fff")
                fig_rank.add_trace(go.Bar(
                    x=sub["CustomerID"].astype(str),
                    y=sub[sort_col],
                    name=pname,
                    marker_color=color,
                    opacity=0.85,
                    hovertemplate=(
                        "<b>KH: %{x}</b><br>"
                        + f"{sort_feature_label}: " + "%{y:,.1f}<br>"
                        + f"Phân khúc: {pname}"
                        + "<extra></extra>"
                    ),
                ))
            prefix = "£" if "Price" in sort_col or "price" in sort_col else ""
            fig_rank.update_layout(
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                font=dict(color="#edf0f5", family="DM Sans"),
                margin=dict(l=40, r=20, t=40, b=40),
                height=420,
                barmode="stack",
                xaxis=dict(showticklabels=False, title="Customer ID", gridcolor="#262c3a", zerolinecolor="#262c3a"),
                yaxis=dict(title=f"{prefix}{sort_feature_label}", gridcolor="#262c3a", zerolinecolor="#262c3a"),
                legend=dict(orientation="h", y=-0.12, x=0.5, xanchor="center"),
            )
            st.plotly_chart(fig_rank, use_container_width=True)

        # ── Bảng chi tiết
        st.markdown("#### Bảng chi tiết")
        display_cols_rank = ["CustomerID", "Phân Khúc"]
        feat_show = [f for f in ["Sum_TotalPrice", "Count_Invoice", "Sum_Quantity", "Mean_UnitPrice", "Count_Stock"] if f in filtered.columns]
        display_cols_rank += feat_show

        table_df = filtered[display_cols_rank].copy()
        rename_map = {"CustomerID": "Customer ID", "Phân Khúc": "Phân Khúc"}
        rename_map.update({f: FEATURE_NAMES_VN.get(f, f) for f in feat_show})
        table_df = table_df.rename(columns=rename_map)

        # Format số
        for col_name in ["Tổng chi tiêu", "Giá trung bình"]:
            if col_name in table_df.columns:
                table_df[col_name] = table_df[col_name].apply(lambda x: f"£{x:,.2f}")
        for col_name in ["Số lần mua", "Tổng số lượng", "Số sản phẩm khác nhau"]:
            if col_name in table_df.columns:
                table_df[col_name] = table_df[col_name].apply(lambda x: f"{x:,.0f}")

        st.dataframe(table_df, use_container_width=True, height=420)

        # Download button
        csv_data = filtered[display_cols_rank].to_csv(index=True).encode("utf-8")
        st.download_button(
            label="⬇️ Tải xuống CSV",
            data=csv_data,
            file_name=f"customer_ranking_{sort_feature_label.replace(' ', '_')}.csv",
            mime="text/csv",
        )

    # ── TAB 2: Danh sách theo phân khúc ──────────────────────────
    with tab_segment:
        st.markdown("#### Khách hàng thuộc từng phân khúc")

        seg_df = merged_df.copy().reset_index()
        seg_df = seg_df.rename(columns={"index": "CustomerID"})
        seg_df["Phân Khúc"] = seg_df[cluster_col].apply(
            lambda x: personas.get(int(x), {}).get("name", f"Cluster {x}") if pd.notna(x) else "N/A"
        )

        # Chọn phân khúc muốn xem
        selected_seg = st.selectbox(
            "Chọn phân khúc",
            options=[personas.get(i, {}).get("name", f"C{i}") for i in sorted(merged_df[cluster_col].dropna().unique().astype(int))],
        )

        seg_filtered = seg_df[seg_df["Phân Khúc"] == selected_seg].copy()

        # Tìm cluster id để lấy màu
        seg_cid = next(
            (i for i in sorted(merged_df[cluster_col].dropna().unique().astype(int))
             if personas.get(i, {}).get("name") == selected_seg), 0
        )
        seg_color = colors.get(seg_cid, "#7c8aff")

        # KPI row cho phân khúc đó
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.metric("Số khách hàng", f"{len(seg_filtered):,}")
        with c2:
            avg_sp = seg_filtered["Sum_TotalPrice"].mean() if "Sum_TotalPrice" in seg_filtered.columns else 0
            st.metric("Chi tiêu TB", f"£{avg_sp:,.0f}")
        with c3:
            avg_inv = seg_filtered["Count_Invoice"].mean() if "Count_Invoice" in seg_filtered.columns else 0
            st.metric("Số mua TB", f"{avg_inv:,.1f} đơn")
        with c4:
            avg_qty = seg_filtered["Sum_Quantity"].mean() if "Sum_Quantity" in seg_filtered.columns else 0
            st.metric("Số lượng TB", f"{avg_qty:,.0f} sp")

        st.markdown("---")

        # Sort option
        col_s1, col_s2 = st.columns([4, 2])
        with col_s1:
            seg_sort_label = st.selectbox(
                "Sắp xếp theo",
                ["Tổng chi tiêu", "Số lần mua", "Tổng số lượng", "Giá trung bình", "Số sản phẩm khác nhau"],
                key="seg_sort",
            )
        with col_s2:
            seg_sort_order = st.radio("Thứ tự", ["Cao → Thấp", "Thấp → Cao"], horizontal=True, key="seg_order")

        seg_sort_col = sort_map[seg_sort_label]
        seg_asc = seg_sort_order == "Thấp → Cao"

        if seg_sort_col in seg_filtered.columns:
            seg_filtered = seg_filtered.sort_values(seg_sort_col, ascending=seg_asc).reset_index(drop=True)
            seg_filtered.index = seg_filtered.index + 1

        # Bảng
        feat_seg = [f for f in ["Sum_TotalPrice", "Count_Invoice", "Sum_Quantity", "Mean_UnitPrice", "Count_Stock"] if f in seg_filtered.columns]
        seg_display = seg_filtered[["CustomerID"] + feat_seg].copy()
        seg_rename = {"CustomerID": "Customer ID"}
        seg_rename.update({f: FEATURE_NAMES_VN.get(f, f) for f in feat_seg})
        seg_display = seg_display.rename(columns=seg_rename)

        for col_name in ["Tổng chi tiêu", "Giá trung bình"]:
            if col_name in seg_display.columns:
                seg_display[col_name] = seg_display[col_name].apply(lambda x: f"£{x:,.2f}")
        for col_name in ["Số lần mua", "Tổng số lượng", "Số sản phẩm khác nhau"]:
            if col_name in seg_display.columns:
                seg_display[col_name] = seg_display[col_name].apply(lambda x: f"{x:,.0f}")

        st.dataframe(seg_display, use_container_width=True, height=500)

        # Download
        csv_seg = seg_filtered[["CustomerID"] + feat_seg].to_csv(index=True).encode("utf-8")
        st.download_button(
            label=f"⬇️ Tải xuống danh sách {selected_seg}",
            data=csv_seg,
            file_name=f"customers_{selected_seg.replace(' ', '_').replace('💎','').replace('🔥','').replace('🌱','').replace('😴','').strip()}.csv",
            mime="text/csv",
        )


# ─────────────────────────────────────────────
# ══ PAGE: TRA CỨU KHÁCH HÀNG ══
# ─────────────────────────────────────────────
elif page == "👤 Tra Cứu Khách Hàng":
    st.markdown("# 👤 Tra Cứu Khách Hàng")
    render_page_hero(
        "Customer Lookup",
        "Xem hồ sơ một khách hàng cụ thể trong ngữ cảnh cluster",
        "Trang tra cứu kết hợp metric cá nhân, radar so sánh và lịch sử giao dịch để đọc nhanh hành vi của một khách hàng.",
        ["Customer profile", "Segment comparison", "Transaction history"],
    )

    if merged_df is None or cluster_col not in merged_df.columns:
        st.warning("Không có dữ liệu cluster.")
        st.stop()

    all_customers = sorted(merged_df.index.astype(str).tolist())

    col_search, col_rand = st.columns([5, 1])
    with col_search:
        selected_customer = st.selectbox("Nhập hoặc chọn Customer ID", all_customers)
    with col_rand:
        st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
        if st.button("🎲 Ngẫu nhiên"):
            selected_customer = np.random.choice(all_customers)

    if selected_customer:
        try:
            row = merged_df.loc[selected_customer]
        except KeyError:
            st.error(f"Không tìm thấy khách hàng: {selected_customer}")
            st.stop()

        cid_int = int(row[cluster_col]) if pd.notna(row[cluster_col]) else -1
        persona = personas.get(cid_int, {"name": "Unknown", "color": "#8894a8", "desc": ""})

        st.markdown(
            f"""<div class='metric-card' style='border-left:4px solid {persona["color"]};margin-bottom:24px'>
                <div style='font-family:Space Mono,monospace;font-size:0.75rem;color:#8894a8;margin-bottom:4px'>CUSTOMER ID</div>
                <div style='font-family:Space Mono,monospace;font-size:1.8rem;font-weight:700'>{selected_customer}</div>
                <div style='margin-top:8px'>
                  <span class='cluster-badge' style='background:{persona["color"]}22;color:{persona["color"]};border:1px solid {persona["color"]}44'>
                    {persona["name"]}
                  </span>
                  <span style='color:#8894a8;font-size:0.82rem;margin-left:10px'>{persona["desc"]}</span>
                </div>
            </div>""",
            unsafe_allow_html=True,
        )

        # Feature values
        feat_cols_display = [f for f in features_df.columns if f in merged_df.columns]
        cust_vals = row[feat_cols_display]
        cluster_means_ref = merged_df.groupby(cluster_col)[feat_cols_display].mean()
        overall_means = merged_df[feat_cols_display].mean()

        col_left, col_right = st.columns(2)
        with col_left:
            st.markdown("#### Chỉ số chính")
            key_metrics = [
                ("Sum_TotalPrice", "Tổng chi tiêu", "£", ""),
                ("Count_Invoice", "Số lần mua", "", " đơn"),
                ("Sum_Quantity", "Tổng số lượng", "", " sp"),
                ("Count_Stock", "Sản phẩm đa dạng", "", " loại"),
                ("Mean_UnitPrice", "Giá TB", "£", ""),
            ]
            for feat, label, prefix, suffix in key_metrics:
                if feat in cust_vals.index:
                    val = cust_vals[feat]
                    mean_val = overall_means[feat]
                    delta_pct = (val - mean_val) / (mean_val + 1e-9) * 100
                    st.metric(label, f"{prefix}{val:,.1f}{suffix}", delta=f"{delta_pct:+.1f}% vs TB")

        with col_right:
            st.markdown("#### So sánh với trung bình phân khúc")
            if cid_int in cluster_means_ref.index:
                cluster_mean_row = cluster_means_ref.loc[cid_int]
                radar_keys = [f for f in RADAR_FEATURES.keys() if f in feat_cols_display]

                # Normalize globally
                global_max_r = merged_df[radar_keys].max()
                global_min_r = merged_df[radar_keys].min()
                cust_norm = (cust_vals[radar_keys] - global_min_r) / (global_max_r - global_min_r + 1e-9)
                cl_norm = (cluster_mean_row[radar_keys] - global_min_r) / (global_max_r - global_min_r + 1e-9)

                cats = [RADAR_FEATURES[f] for f in radar_keys]
                fig_cust_radar = go.Figure()
                for norm_vals, trace_name, color, opacity in [
                    (cust_norm.tolist(), f"KH {selected_customer}", persona["color"], 0.3),
                    (cl_norm.tolist(), "TB Phân khúc", "#8894a8", 0.1),
                ]:
                    h = color.lstrip("#")
                    rc, gc, bc = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
                    vals = norm_vals + norm_vals[:1]
                    c_list = cats + cats[:1]
                    fig_cust_radar.add_trace(go.Scatterpolar(
                        r=vals, theta=c_list, name=trace_name,
                        fill="toself", fillcolor=f"rgba({rc},{gc},{bc},{opacity})",
                        line=dict(color=color, width=2),
                    ))

                fig_cust_radar.update_layout(
                    **PLOTLY_LAYOUT, height=340,
                    polar=dict(
                        bgcolor="rgba(0,0,0,0)",
                        radialaxis=dict(visible=True, range=[0, 1], gridcolor="#262c3a", color="#8894a8"),
                        angularaxis=dict(gridcolor="#262c3a", color="#edf0f5"),
                    ),
                    legend=dict(orientation="h", y=-0.15, x=0.5, xanchor="center"),
                )
                st.plotly_chart(fig_cust_radar, use_container_width=True)

        # All features bar
        st.markdown("#### Tất cả features (so với TB phân khúc)")
        if cid_int in cluster_means_ref.index:
            compare_df = pd.DataFrame({
                "Feature": [FEATURE_NAMES_VN.get(f, f) for f in feat_cols_display],
                "Khách hàng": cust_vals[feat_cols_display].values,
                "TB Phân khúc": cluster_means_ref.loc[cid_int, feat_cols_display].values,
            })
            # Normalize for display
            max_vals = compare_df[["Khách hàng", "TB Phân khúc"]].max(axis=1)
            compare_norm = compare_df.copy()
            compare_norm["Khách hàng"] = compare_df["Khách hàng"] / (max_vals + 1e-9)
            compare_norm["TB Phân khúc"] = compare_df["TB Phân khúc"] / (max_vals + 1e-9)

            fig_cmp = go.Figure()
            fig_cmp.add_trace(go.Bar(name=f"KH {selected_customer}", x=compare_norm["Feature"],
                                     y=compare_norm["Khách hàng"], marker_color=persona["color"], opacity=0.85))
            fig_cmp.add_trace(go.Bar(name="TB Phân khúc", x=compare_norm["Feature"],
                                     y=compare_norm["TB Phân khúc"], marker_color="#8894a8", opacity=0.6))
            fig_cmp.update_layout(**PLOTLY_LAYOUT, height=340, barmode="group",
                                  xaxis_tickangle=-35, legend=dict(orientation="h", y=-0.25))
            st.plotly_chart(fig_cmp, use_container_width=True)

        # Transaction history
        if cleaned_df is not None:
            st.markdown("#### Lịch sử giao dịch")
            cust_txns = cleaned_df[cleaned_df["CustomerID"].astype(str) == str(selected_customer)]

            if len(cust_txns) > 0:
                st.markdown(f"**{len(cust_txns):,} giao dịch** · Tổng: £{cust_txns['TotalPrice'].sum():,.2f}")

                display_cols = [
                    c for c in ["InvoiceDate", "InvoiceNo", "Description", "Quantity", "UnitPrice", "TotalPrice"]
                    if c in cust_txns.columns
                ]

                st.dataframe(
                    cust_txns[display_cols].sort_values("InvoiceDate", ascending=False),
                    use_container_width=True,
                    hide_index=True,
                )
            else:
                st.info("Không tìm thấy lịch sử giao dịch cho khách hàng này.")

# ─────────────────────────────────────────────
# ══ PAGE: MÔ PHỎNG KHÁCH HÀNG ══
# ─────────────────────────────────────────────
elif page == "🧪 Mô Phỏng Khách Hàng":
    st.markdown("# 🧪 Mô Phỏng Khách Hàng")
    render_page_hero(
        "What-If Simulator",
        "Một khách hàng giả định sẽ thuộc phân khúc nào?",
        "Điều chỉnh chỉ số hành vi mua hàng để xem khách hàng giả định gần với phân khúc nào nhất. Công cụ minh họa nhanh dùng khoảng cách chuẩn hóa tới tâm cụm — không phải mô hình K-Means chính thức đã huấn luyện trong notebook.",
        ["Interactive sliders", "Nearest-centroid", "Radar so sánh"],
    )

    if merged_df is None or cluster_col not in merged_df.columns:
        st.warning("Không có dữ liệu cluster.")
        st.stop()

    sim_features = ["Sum_Quantity", "Mean_UnitPrice", "Sum_TotalPrice", "Count_Invoice", "Count_Stock"]
    sim_features = [f for f in sim_features if f in merged_df.columns]

    st.markdown("### Nhập chỉ số khách hàng giả định")
    sim_cols = st.columns(len(sim_features))
    sim_input = {}
    for col, feat in zip(sim_cols, sim_features):
        with col:
            series = merged_df[feat].dropna()
            lo, hi = float(series.quantile(0.01)), float(series.quantile(0.99))
            default = float(series.median())
            sim_input[feat] = st.slider(
                FEATURE_NAMES_VN.get(feat, feat),
                min_value=lo, max_value=hi, value=default,
                key=f"sim_{feat}",
            )

    # Chuẩn hóa bằng mean/std của chính bộ dữ liệu (xấp xỉ scaler đã dùng trong notebook)
    means = merged_df[sim_features].mean()
    stds = merged_df[sim_features].std().replace(0, 1)
    input_z = pd.Series({f: (sim_input[f] - means[f]) / stds[f] for f in sim_features})

    cluster_means_raw = merged_df.groupby(cluster_col)[sim_features].mean()
    cluster_means_z = (cluster_means_raw - means) / stds

    distances = ((cluster_means_z - input_z) ** 2).sum(axis=1).pow(0.5)
    nearest_cid = int(distances.idxmin())
    persona_pred = personas.get(nearest_cid, {"name": "Unknown", "color": "#8894a8", "desc": ""})

    st.markdown("---")
    st.markdown(
        f"""<div class='metric-card' style='border-left:4px solid {persona_pred["color"]};margin-bottom:16px'>
            <div style='font-family:Space Mono,monospace;font-size:0.75rem;color:#8894a8;margin-bottom:4px'>PHÂN KHÚC GẦN NHẤT</div>
            <div class='metric-value' style='color:{persona_pred["color"]};font-size:1.6rem'>{persona_pred["name"]}</div>
            <div class='metric-label'>{persona_pred["desc"]}</div>
        </div>""",
        unsafe_allow_html=True,
    )

    dist_df = pd.DataFrame({
        "Phân khúc": [personas.get(int(i), {}).get("name", f"C{i}") for i in distances.index],
        "Khoảng cách chuẩn hóa": distances.values,
    }).sort_values("Khoảng cách chuẩn hóa")
    st.dataframe(dist_df, use_container_width=True, hide_index=True)

    st.markdown("### So sánh với phân khúc dự đoán")
    global_min_s = merged_df[sim_features].min()
    global_max_s = merged_df[sim_features].max()

    input_norm = pd.Series({
        f: (sim_input[f] - global_min_s[f]) / (global_max_s[f] - global_min_s[f] + 1e-9)
        for f in sim_features
    })
    cluster_mean_raw = cluster_means_raw.loc[nearest_cid]
    cluster_norm = (cluster_mean_raw - global_min_s) / (global_max_s - global_min_s + 1e-9)

    cats_sim = [FEATURE_NAMES_VN.get(f, f) for f in sim_features]
    fig_sim_radar = go.Figure()
    for norm_vals, trace_name, color, opacity in [
        (input_norm.tolist(), "Khách hàng giả định", persona_pred["color"], 0.35),
        (cluster_norm.tolist(), f"TB {persona_pred['name']}", "#8894a8", 0.1),
    ]:
        h = color.lstrip("#")
        rc, gc, bc = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
        vals = norm_vals + norm_vals[:1]
        c_list = cats_sim + cats_sim[:1]
        fig_sim_radar.add_trace(go.Scatterpolar(
            r=vals, theta=c_list, name=trace_name,
            fill="toself", fillcolor=f"rgba({rc},{gc},{bc},{opacity})",
            line=dict(color=color, width=2),
        ))
    fig_sim_radar.update_layout(
        **PLOTLY_LAYOUT, height=400,
        polar=dict(
            bgcolor="rgba(0,0,0,0)",
            radialaxis=dict(visible=True, range=[0, 1], gridcolor="#262c3a", color="#8894a8"),
            angularaxis=dict(gridcolor="#262c3a", color="#edf0f5"),
        ),
        legend=dict(orientation="h", y=-0.15, x=0.5, xanchor="center"),
    )
    st.plotly_chart(fig_sim_radar, use_container_width=True)

    st.caption(
        "⚠️ Công cụ minh họa dùng khoảng cách Euclidean chuẩn hóa tới tâm cụm trên feature gốc "
        "(không qua PCA, không dùng đúng Box-Cox/StandardScaler đã huấn luyện trong notebook). "
        "Kết quả có thể lệch nhẹ so với mô hình K-Means chính thức — phù hợp để minh họa trực quan, "
        "không dùng làm căn cứ quyết định kinh doanh chính thức."
    )


# ─────────────────────────────────────────────
# ══ PAGE: CẬP NHẬT DỮ LIỆU MỚI ══
# ─────────────────────────────────────────────
elif page == "📤 Cập Nhật Dữ Liệu Mới":
    st.markdown("# 📤 Cập Nhật Dữ Liệu Mới")
    render_page_hero(
        "Data Refresh",
        "Gộp dữ liệu giao dịch mới và tính lại toàn bộ phân khúc",
        "Tải lên file CSV giao dịch mới theo đúng cấu trúc dữ liệu gốc. Hệ thống sẽ làm sạch, gộp với dữ liệu "
        "hiện có, và tính lại toàn bộ pipeline — đặc trưng khách hàng, Box-Cox, PCA và phân cụm K-Means "
        "(k=3, k=4) — ngay trên dashboard, không cần chạy lại notebook.",
        ["Upload CSV", "Auto-merge", "Recompute pipeline"],
    )

    st.info(
        "**File CSV cần có đúng các cột:** `InvoiceNo`, `StockCode`, `Description`, `Quantity`, "
        "`InvoiceDate`, `UnitPrice`, `CustomerID`, `Country` — giống cấu trúc dữ liệu gốc (Online Retail Dataset). "
        "Hệ thống sẽ tự động loại hóa đơn hủy, lọc thị trường UK, và loại bản ghi thiếu CustomerID trước khi gộp."
    )

    if st.session_state.get("recomputed"):
        rec = st.session_state["recomputed"]
        if st.session_state.get("just_restored"):
            st.info("🔄 Đã tự động khôi phục dữ liệu đã gộp từ lần cập nhật gần nhất (lưu trên đĩa).")
            st.session_state["just_restored"] = False

        new_label = f"+{rec['n_new_customers']} khách hàng mới, " if rec.get("n_new_customers") is not None else ""
        st.success(
            f"🟢 Dashboard đang hiển thị **dữ liệu đã gộp**: {new_label}"
            f"{rec['n_total_customers']} khách hàng tổng cộng. Tất cả các trang khác đã tự động cập nhật theo dữ liệu này. "
            f"Dữ liệu này **đã được lưu trên đĩa** nên sẽ không mất khi bạn tải lại trang (F5)."
        )
        col_reset1, col_reset2 = st.columns(2)
        with col_reset1:
            if st.button("↩️ Khôi phục dữ liệu gốc (chỉ phiên này)"):
                st.session_state["recomputed"] = None
                st.rerun()
        with col_reset2:
            if st.button("🗑️ Xóa hẳn dữ liệu đã lưu trên đĩa"):
                delete_merged_from_disk()
                st.session_state["recomputed"] = None
                st.rerun()
        st.markdown("---")

    uploaded_file = st.file_uploader("Chọn file CSV giao dịch mới", type=["csv"])

    if uploaded_file is not None:
        try:
            new_raw = pd.read_csv(uploaded_file, encoding="latin1")
        except Exception:
            uploaded_file.seek(0)
            new_raw = pd.read_csv(uploaded_file)

        st.markdown("### Xem trước dữ liệu tải lên")
        st.dataframe(new_raw.head(10), use_container_width=True)
        st.caption(f"{len(new_raw):,} dòng giao dịch trong file tải lên.")

        if st.button("🔄 Gộp & Tính Lại Toàn Bộ Phân Khúc", type="primary"):
            with st.spinner("Đang làm sạch, gộp dữ liệu và tính lại pipeline — có thể mất vài giây..."):
                try:
                    new_clean = clean_new_upload(new_raw)
                except ValueError as e:
                    st.error(f"❌ {e}")
                    st.stop()

                if len(new_clean) == 0:
                    st.warning(
                        "Sau khi làm sạch, không còn dòng giao dịch hợp lệ nào. "
                        "Kiểm tra lại cột `Country` (phải là 'United Kingdom') và `CustomerID` (không được để trống)."
                    )
                    st.stop()

                base_clean = data["cleaned"].copy() if data["cleaned"] is not None else pd.DataFrame(columns=new_clean.columns)
                if "CustomerID" in base_clean.columns and len(base_clean) > 0:
                    base_clean["CustomerID"] = normalize_customer_id(base_clean["CustomerID"])
                combined_clean = pd.concat([base_clean, new_clean], ignore_index=True).drop_duplicates()

                old_customers = set(normalize_customer_id(pd.Series(features_df.index))) if features_df is not None else set()

                new_features = build_customer_features(combined_clean)
                added_customers = set(new_features.index) - old_customers

                new_scaled, pca_coords, evr, cluster_results = transform_and_cluster(new_features, k_list=(3, 4))

                st.session_state["recomputed"] = {
                    "cleaned": combined_clean,
                    "features": new_features,
                    "scaled": new_scaled,
                    "clusters_k3": cluster_results[3],
                    "clusters_k4": cluster_results[4],
                    "n_new_customers": len(added_customers),
                    "n_total_customers": len(new_features),
                }
                saved_ok, save_err = save_merged_to_disk(st.session_state["recomputed"])

            st.success(
                f"✅ Đã gộp và tính lại thành công! **{len(added_customers)} khách hàng mới**, "
                f"tổng cộng **{len(new_features)} khách hàng**. PCA giữ lại {len(evr)} thành phần chính "
                f"({np.sum(evr):.1%} phương sai giải thích)."
            )
            if saved_ok:
                st.caption("💾 Đã lưu kết quả xuống `data/processed/*_merged.csv` — sẽ không mất khi tải lại trang.")
            else:
                st.warning(
                    f"⚠️ Không thể lưu xuống đĩa ({save_err}). Dữ liệu vẫn dùng được trong phiên làm việc hiện tại, "
                    "nhưng sẽ mất nếu tải lại trang (F5). Hãy dùng nút tải CSV bên dưới để lưu thủ công."
                )

            st.markdown("### So sánh phân bố cụm trước và sau khi gộp (k=4)")
            old_k4 = data["clusters_k4"]
            new_k4 = cluster_results[4]
            old_dist = (old_k4["Cluster"].value_counts(normalize=True).sort_index() * 100) if old_k4 is not None else None
            new_dist = new_k4["Cluster"].value_counts(normalize=True).sort_index() * 100

            comp_fig = go.Figure()
            if old_dist is not None:
                comp_fig.add_trace(go.Bar(
                    x=[f"Cụm {i}" for i in old_dist.index], y=old_dist.values,
                    name="Trước khi gộp", marker_color="#7b8794",
                ))
            comp_fig.add_trace(go.Bar(
                x=[f"Cụm {i}" for i in new_dist.index], y=new_dist.values,
                name="Sau khi gộp", marker_color="#7c8aff",
            ))
            comp_fig.update_layout(
                **PLOTLY_LAYOUT, height=380, barmode="group",
                yaxis_title="Tỷ lệ khách hàng (%)",
                legend=dict(orientation="h", y=-0.15, x=0.5, xanchor="center"),
            )
            st.plotly_chart(comp_fig, use_container_width=True)

            st.info(
                "💡 Chuyển sang các trang khác (Tổng Quan, Khám Phá Phân Khúc, Xếp Hạng Khách Hàng...) "
                "để xem toàn bộ dashboard đã cập nhật theo dữ liệu mới."
            )

            st.markdown("### Tải kết quả đã tính lại")
            dl_col1, dl_col2 = st.columns(2)
            with dl_col1:
                st.download_button(
                    "⬇️ Tải customer_features_updated.csv",
                    new_features.to_csv(index=True).encode("utf-8"),
                    file_name="customer_features_updated.csv",
                    mime="text/csv",
                )
            with dl_col2:
                st.download_button(
                    "⬇️ Tải customer_clusters_k4_updated.csv",
                    cluster_results[4].to_csv(index=False).encode("utf-8"),
                    file_name="customer_clusters_k4_updated.csv",
                    mime="text/csv",
                )

    st.markdown("---")
    st.caption(
        "⚠️ Lưu ý: việc tính lại sẽ **huấn luyện lại từ đầu** Box-Cox, StandardScaler, PCA và K-Means trên toàn bộ "
        "dữ liệu đã gộp (không phải chỉ dự đoán nhãn cho khách hàng mới bằng mô hình cũ). Vì vậy tâm cụm, ranh giới "
        "phân khúc, và thậm chí thứ tự đánh số cụm (0/1/2/3) có thể thay đổi nhẹ so với kết quả gốc trong luận văn — "
        "đây là hành vi đúng theo thiết kế khi \"tính lại\" trên dữ liệu lớn hơn, không phải lỗi.\n\n"
        "Dữ liệu đã gộp được lưu vào các file riêng `*_merged.csv` trong `data/processed/` — "
        "**không ghi đè** lên các file gốc (`cleaned_uk_data.csv`, `customer_features.csv`...), nên bạn luôn có thể "
        "quay lại baseline gốc bằng nút \"Khôi phục dữ liệu gốc\". Vì được lưu trên đĩa, dữ liệu đã gộp sẽ **không mất** "
        "khi tải lại trang (F5) — chỉ mất khi bạn bấm \"Xóa hẳn dữ liệu đã lưu trên đĩa\", hoặc khi deploy trên các nền "
        "tảng có ổ đĩa tạm thời (ephemeral filesystem) như Streamlit Community Cloud, nơi dữ liệu có thể mất sau khi app khởi động lại."
    )


# ─────────────────────────────────────────────
# ══ PAGE: XUẤT BÁO CÁO TỔNG HỢP ══
# ─────────────────────────────────────────────
elif page == "📄 Xuất Báo Cáo Tổng Hợp":
    st.markdown("# 📄 Xuất Báo Cáo Tổng Hợp")
    render_page_hero(
        "Report Export",
        "Xuất báo cáo/dashboard tổng hợp để trình bày hoặc phân tích thêm",
        "Chọn định dạng phù hợp: PDF để trình bày hoặc gửi cho ban lãnh đạo, Excel dạng dashboard có "
        "bộ lọc nhanh và biểu đồ để tự phân tích sâu hơn, hoặc ảnh PNG/JPEG để dán nhanh vào slide/tài liệu "
        "— dựa trên dữ liệu hiện tại của dashboard.",
        ["PDF Report", "Excel Dashboard", "Ảnh PNG/JPEG"],
    )

    tab_pdf, tab_excel, tab_image = st.tabs(["📄 PDF Report", "📊 Excel Dashboard", "🖼️ Ảnh Dashboard"])
    data_label = "dữ liệu đã gộp" if st.session_state.get("recomputed") else "dữ liệu gốc"

    with tab_pdf:
        if not PDF_EXPORT_AVAILABLE:
            st.error(
                "⚠️ Thiếu thư viện để xuất PDF. Hãy cài đặt thêm bằng lệnh:\n\n"
                "`pip install matplotlib reportlab`\n\n"
                f"Chi tiết lỗi: `{_PDF_IMPORT_ERROR_MSG}`"
            )
        else:
            st.info(
                f"Báo cáo sẽ được tạo từ **{data_label}** hiện đang hiển thị trên dashboard "
                f"({len(features_df):,} khách hàng)."
            )
            st.caption(
                "Gồm: tổng quan dữ liệu, phân bố khách hàng theo phân khúc (k=3, k=4), radar chart đặc trưng "
                "hành vi từng nhóm, và đề xuất chiến lược marketing."
            )

            if st.button("🖨️ Tạo Báo Cáo PDF", type="primary"):
                with st.spinner("Đang tạo báo cáo PDF — có thể mất vài giây..."):
                    try:
                        pdf_bytes = build_pdf_report(
                            cleaned_df, features_df, data["clusters_k3"], data["clusters_k4"], data_label,
                        )
                        st.session_state["pdf_report_bytes"] = pdf_bytes
                        st.success("✅ Đã tạo báo cáo thành công! Bấm nút bên dưới để tải xuống.")
                    except Exception as e:
                        st.error(f"❌ Có lỗi khi tạo báo cáo: {e}")

            if st.session_state.get("pdf_report_bytes"):
                st.download_button(
                    "⬇️ Tải Báo Cáo PDF",
                    data=st.session_state["pdf_report_bytes"],
                    file_name=f"bao_cao_phan_khuc_khach_hang_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
                    mime="application/pdf",
                    type="primary",
                )

            st.caption(
                "💡 Dùng font DejaVu Sans (đóng gói sẵn trong matplotlib) nên hiển thị đầy đủ dấu tiếng Việt, "
                "không cần cài thêm font riêng."
            )

    with tab_excel:
        if not EXCEL_EXPORT_AVAILABLE:
            st.error(
                "⚠️ Thiếu thư viện để xuất Excel. Hãy cài đặt thêm bằng lệnh:\n\n"
                "`pip install openpyxl`\n\n"
                f"Chi tiết lỗi: `{_XLSX_IMPORT_ERROR_MSG}`"
            )
        else:
            st.info(
                f"File Excel sẽ được tạo từ **{data_label}** hiện đang hiển thị trên dashboard "
                f"({len(features_df):,} khách hàng)."
            )
            st.caption(
                "Gồm 3 sheet: **Dashboard** (KPI + ô lọc nhanh theo phân khúc + biểu đồ), "
                "**Dữ Liệu Khách Hàng** (bảng đầy đủ có bộ lọc từng cột — AutoFilter), "
                "và **So Sánh Cụm** (bảng 16 đặc trưng, tô màu theo giá trị)."
            )
            st.warning(
                "⚠️ Lưu ý: ô lọc trên sheet Dashboard là **Data Validation dropdown kết hợp công thức** "
                "(SUMIF/COUNTIF/AVERAGEIF), không phải Slicer gốc của Excel/Power BI — hiệu quả lọc số liệu "
                "tương tự, nhưng không có giao diện nút bấm dạng \"chip\" như Slicer thật.",
                icon="ℹ️",
            )

            if st.button("📊 Tạo Excel Dashboard", type="primary"):
                with st.spinner("Đang tạo file Excel — có thể mất vài giây..."):
                    try:
                        xlsx_bytes = build_excel_dashboard(
                            cleaned_df, features_df, data["clusters_k3"], data["clusters_k4"], data_label,
                        )
                        st.session_state["xlsx_report_bytes"] = xlsx_bytes
                        st.success("✅ Đã tạo file Excel thành công! Bấm nút bên dưới để tải xuống.")
                    except Exception as e:
                        st.error(f"❌ Có lỗi khi tạo file Excel: {e}")

            if st.session_state.get("xlsx_report_bytes"):
                st.download_button(
                    "⬇️ Tải Excel Dashboard",
                    data=st.session_state["xlsx_report_bytes"],
                    file_name=f"dashboard_phan_khuc_khach_hang_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                )

    with tab_image:
        if not PDF_EXPORT_AVAILABLE:  # dùng chung matplotlib với phần PDF
            st.error(
                "⚠️ Thiếu thư viện matplotlib để xuất ảnh. Hãy cài đặt bằng lệnh:\n\n"
                "`pip install matplotlib`\n\n"
                f"Chi tiết lỗi: `{_PDF_IMPORT_ERROR_MSG}`"
            )
        else:
            st.info(
                f"Ảnh dashboard sẽ được tạo từ **{data_label}** hiện đang hiển thị trên dashboard "
                f"({len(features_df):,} khách hàng)."
            )
            st.caption(
                "Ghép các \"widget\": thẻ chỉ số KPI, biểu đồ tròn phân bố phân khúc, biểu đồ cột so sánh "
                "chi tiêu, và radar chart đặc trưng hành vi — thành **một ảnh duy nhất**, phù hợp để dán vào "
                "slide, tài liệu, hoặc chia sẻ nhanh qua chat."
            )

            img_format = st.radio("Định dạng ảnh", ["PNG (nét hơn, khuyến nghị)", "JPEG (nhẹ hơn)"],
                                   horizontal=True, label_visibility="collapsed")
            fmt_code = "png" if img_format.startswith("PNG") else "jpeg"

            if st.button("🖼️ Tạo Ảnh Dashboard", type="primary"):
                with st.spinner("Đang dựng ảnh dashboard — có thể mất vài giây..."):
                    try:
                        img_bytes = build_dashboard_image(
                            cleaned_df, features_df, data["clusters_k3"], data["clusters_k4"],
                            data_label, img_format=fmt_code,
                        )
                        st.session_state["dashboard_img_bytes"] = img_bytes
                        st.session_state["dashboard_img_fmt"] = fmt_code
                        st.success("✅ Đã tạo ảnh thành công!")
                    except Exception as e:
                        st.error(f"❌ Có lỗi khi tạo ảnh: {e}")

            if st.session_state.get("dashboard_img_bytes"):
                try:
                    st.image(st.session_state["dashboard_img_bytes"], use_container_width=True)
                except TypeError:
                    # Streamlit phiên bản cũ hơn chưa hỗ trợ use_container_width cho st.image,
                    # chỉ hỗ trợ use_column_width (tên tham số cũ).
                    st.image(st.session_state["dashboard_img_bytes"], use_column_width=True)
                ext = st.session_state.get("dashboard_img_fmt", "png")
                mime = "image/png" if ext == "png" else "image/jpeg"
                st.download_button(
                    f"⬇️ Tải Ảnh Dashboard (.{ext})",
                    data=st.session_state["dashboard_img_bytes"],
                    file_name=f"dashboard_phan_khuc_khach_hang_{datetime.now().strftime('%Y%m%d_%H%M')}.{ext}",
                    mime=mime,
                    type="primary",
                )

    st.markdown("---")
    st.caption(
        "Nội dung xuất luôn phản ánh đúng dữ liệu đang hiển thị trên dashboard tại thời điểm bấm nút tạo — "
        "nếu vừa gộp dữ liệu mới ở trang trước, hãy tạo lại để cập nhật số liệu mới nhất."
    )